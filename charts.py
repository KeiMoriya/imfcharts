'''
2025-10-05

Generate charts.
'''

from __future__ import annotations

import os
import sys
import itertools
import warnings
from collections import OrderedDict

import numpy as np
import pandas as pd

import matplotlib
import matplotlib.colors as mplcolors
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.ticker import FuncFormatter, MaxNLocator, ScalarFormatter
from matplotlib.patches import Patch

def _parse_cols(cols):
    '''
    Utility function to parse input columns.
    '''

    # Parse list ccols to get which columns to use.
    # These will be saved in use_cols.
    use_cols = []
    
    # If None is given, don't add anything
    if cols is None:
        return use_cols

    # If a str was given, check whether such a column exists and add it.
    elif type(cols) == str:
        use_cols.append(cols)
        
    # If an iterable is given, check whether valid and add.
    else:
        try:
            for col in cols:
                if type(col) == str:
                    use_cols.append(col)
                else:
                    print('cols must be given as str or iterable of str,')
                    print('given col of type ' + str(type(col)) + ':')
                    print(col)
                    raise ValueError
        except TypeError:
            error = '_parse_cols: cols must be given as str or iterable of str, '
            error += 'given type of ' + str(type(cols)) + ':\n'
            error += cols
            raise ValueError(error)

    # Remove any duplicate columns
    # (using a dict in Python 3.7 onwards will guarantee original ordering)
    use_cols = list(dict.fromkeys(use_cols))

    return use_cols

def guess_freq(df):
    '''
    Guess frequency of a series given a DataFrame.
    Checks differences in days between index and tries go guess.
    '''

    # If df is None, return '?', just a holder
    if df is None:
        return '?'

    # Not enough data to check differences.
    # Return daily freq.
    if len(df) < 2:
        return 'D'
    else:
        idx_types = {type(idx) for idx in df.index}
        if idx_types in [{pd.Timestamp}, {pd.Period}]:
            if (df.index[1] - df.index[0]).days < 6:
                return 'D'
            elif (df.index[1] - df.index[0]).days < 10:
                return 'W'
            elif (df.index[1] - df.index[0]).days < 32:
                return 'M'
            elif (df.index[1] - df.index[0]).days < 93:
                return 'Q'
            elif (df.index[1] - df.index[0]).days < 370:
                return 'A'
            else:
                return 'D'
        else:
            # If not time series, return '?'
            return '?'

class Chart:
    '''
    Class that generates and modifies charts.

    Provides easy way to generate custom charts using custom styles
    and contains internal matplotlib.Figure object for further manipulation.
    Able to save internal figure object.
    '''

    _SPEC_SECTIONS = {
        "data",
        "marks",
        "text",
        "axes",
        "legend",
        "annotations",
        "layout",
        "watermark",
        "debug",
    }
    _MARK_TYPES = {"line", "bar", "area"}
    _MARK_AXES = {"left", "right"}
    _LINE_STYLE_KEYS = {
        "marker",
        "markersize",
        "markerfacecolor",
        "markeredgecolor",
        "markeredgewidth",
        "linewidth",
        "linestyle",
        "color",
        "drawstyle",
        "dashes",
        "dash_capstyle",
        "legend",
    }
    _BAR_STYLE_KEYS = {
        "color",
        "hatchcolor",
        "hatch",
        "hatchwidth",
        "edgecolor",
        "offset",
        "width",
        "legend",
        "barcolors",
    }
    _AREA_STYLE_KEYS = {
        "color",
        "hatchcolor",
        "hatch",
        "hatchwidth",
        "edgecolor",
        "alpha",
        "legend",
    }
    _STYLE_KEYS_BY_MARK = {
        "line": _LINE_STYLE_KEYS,
        "bar": _BAR_STYLE_KEYS,
        "area": _AREA_STYLE_KEYS,
    }
    _ANNOTATION_TYPES = {"hlines", "vlines", "hrects", "vrects", "fills", "texts", "arrows"}

    @staticmethod
    def _copy_data(data):
        if isinstance(data, pd.DataFrame):
            return data.copy()
        return data

    @staticmethod
    def _as_list(value):
        if value is None:
            return []
        if isinstance(value, str):
            return [value]
        try:
            return list(value)
        except TypeError:
            return [value]

    @staticmethod
    def _merge_attrs(base_attrs, cols, style):
        attrs = {}
        if base_attrs:
            attrs.update({key: value.copy() for key, value in base_attrs.items()})
        if style:
            for col in Chart._as_list(cols):
                attrs.setdefault(col, {})
                attrs[col].update(style)
        return attrs or None

    @classmethod
    def _validate_mark_style(cls, mark_type, style):
        if style is None:
            return
        if not isinstance(style, dict):
            raise TypeError(f"mark style must be a dict, got {type(style)!r}")
        allowed = cls._STYLE_KEYS_BY_MARK[mark_type]
        unknown = set(style) - allowed
        if unknown:
            raise ValueError(
                f"Unknown {mark_type} style keys: {sorted(unknown)}. "
                f"Allowed keys are: {sorted(allowed)}"
            )

    @classmethod
    def _validate_attrs(cls, mark_type, attrs):
        if attrs is None:
            return
        if not isinstance(attrs, dict):
            raise TypeError(f"attrs must be a dict, got {type(attrs)!r}")
        for series, style in attrs.items():
            try:
                cls._validate_mark_style(mark_type, style)
            except Exception as exc:
                raise type(exc)(f"Invalid attrs for series {series!r}: {exc}") from exc

    @classmethod
    def validate_spec(cls, spec):
        """
        Validate a declarative chart specification.

        The spec format is intentionally close to JSON/YAML so external systems
        can generate chart requests without knowing the long Chart constructor
        signature.
        """

        if not isinstance(spec, dict):
            raise TypeError(f"spec must be a dict, got {type(spec)!r}")

        unknown_sections = set(spec) - cls._SPEC_SECTIONS
        if unknown_sections:
            raise ValueError(f"Unknown spec sections: {sorted(unknown_sections)}")

        marks = spec.get("marks", [])
        if marks is None:
            marks = []
        if not isinstance(marks, list):
            raise TypeError("spec['marks'] must be a list")

        for i, mark in enumerate(marks):
            if not isinstance(mark, dict):
                raise TypeError(f"mark {i} must be a dict")

            mark_type = mark.get("type")
            if mark_type not in cls._MARK_TYPES:
                raise ValueError(f"mark {i} has unsupported type {mark_type!r}; use one of {sorted(cls._MARK_TYPES)}")

            if "cols" not in mark:
                raise ValueError(f"mark {i} is missing required key 'cols'")

            axis = mark.get("axis", "left")
            if axis not in cls._MARK_AXES:
                raise ValueError(f"mark {i} has unsupported axis {axis!r}; use 'left' or 'right'")

            allowed_mark_keys = {
                "type",
                "cols",
                "axis",
                "style",
                "attrs",
                "data",
                "indexcol",
                "stack",
                "colors",
                "linewidth",
                "linebreaks",
                "drawstyle",
                "total_barwidth",
                "edgecolor",
                "alpha",
                "xrange",
                "margins",
                "xformat",
                "yrange",
                "ryrange",
                "debug",
            }
            unknown_mark_keys = set(mark) - allowed_mark_keys
            if unknown_mark_keys:
                raise ValueError(f"mark {i} has unknown keys: {sorted(unknown_mark_keys)}")

            cls._validate_mark_style(mark_type, mark.get("style"))
            cls._validate_attrs(mark_type, mark.get("attrs"))

        annotations = spec.get("annotations", {})
        if annotations is None:
            annotations = {}
        if not isinstance(annotations, dict):
            raise TypeError("spec['annotations'] must be a dict")
        unknown_annotations = set(annotations) - cls._ANNOTATION_TYPES
        if unknown_annotations:
            raise ValueError(f"Unknown annotation groups: {sorted(unknown_annotations)}")
        for key, values in annotations.items():
            if values is not None and not isinstance(values, list):
                raise TypeError(f"annotations[{key!r}] must be a list of dicts")

        return True

    @classmethod
    def from_spec(cls, spec, data=None):
        """
        Build a Chart from a JSON-like specification.

        Data can be passed separately with the `data` argument, which is the
        preferred integration path. If spec['data'] is a DataFrame, it is also
        accepted for Python-only usage.
        """

        cls.validate_spec(spec)

        data_spec = spec.get("data", {}) or {}
        if isinstance(data_spec, pd.DataFrame):
            chart_data = data_spec
            data_options = {}
        elif isinstance(data_spec, dict):
            chart_data = data if data is not None else data_spec.get("frame")
            data_options = data_spec
        else:
            raise TypeError("spec['data'] must be a dict or pandas DataFrame")

        chart_data = cls._copy_data(chart_data)
        original_data = cls._copy_data(chart_data)
        indexcol = data_options.get("indexcol")

        text = spec.get("text", {}) or {}
        axes = spec.get("axes", {}) or {}
        xaxis = axes.get("x", {}) or {}
        yaxis = axes.get("y", {}) or {}
        right_yaxis = axes.get("right_y", axes.get("righty", {})) or {}
        legend = spec.get("legend", {}) or {}
        annotations = spec.get("annotations", {}) or {}
        layout = spec.get("layout", {}) or {}
        watermark = spec.get("watermark", {}) or {}

        chart = cls(
            data=chart_data,
            indexcol=indexcol,
            title=text.get("title"),
            subtitle=text.get("subtitle"),
            xtitle=text.get("xtitle", ""),
            ytitle=text.get("ytitle", ""),
            xrange=xaxis.get("range"),
            xformat=xaxis.get("format", "auto"),
            margins=xaxis.get("margins", "auto"),
            yrange=yaxis.get("range"),
            ryrange=right_yaxis.get("range"),
            width=layout.get("width", 10),
            height=layout.get("height", 6),
            topaxis=layout.get("topaxis", "left"),
            legend=legend.get("show", True),
            ncol_legend=legend.get("ncol", legend.get("ncol_legend", 1)),
            legend_left=legend.get("left", legend.get("legend_left", 0.04)),
            legend_bottom=legend.get("bottom", legend.get("legend_bottom", 0.85)),
            legend_width=legend.get("width", legend.get("legend_width", 0.70)),
            legend_height=legend.get("height", legend.get("legend_height", 0.15)),
            legend_header=legend.get("header", legend.get("legend_header", "")),
            watermark=watermark.get("text") if isinstance(watermark, dict) else None,
            wmx=watermark.get("x") if isinstance(watermark, dict) else None,
            wmy=watermark.get("y") if isinstance(watermark, dict) else None,
            wmsize=watermark.get("size") if isinstance(watermark, dict) else None,
            wmfont=watermark.get("font") if isinstance(watermark, dict) else None,
            wmcolor=watermark.get("color") if isinstance(watermark, dict) else None,
            wmalpha=watermark.get("alpha") if isinstance(watermark, dict) else None,
            wmangle=watermark.get("angle") if isinstance(watermark, dict) else None,
            wmzorder=watermark.get("zorder") if isinstance(watermark, dict) else None,
            debug=spec.get("debug", False),
        )

        for mark in spec.get("marks", []) or []:
            mark_type = mark["type"]
            cols = mark["cols"]
            attrs = cls._merge_attrs(mark.get("attrs"), cols, mark.get("style"))
            mark_data = cls._copy_data(mark.get("data", original_data))
            mark_indexcol = mark.get("indexcol", indexcol)

            common = {
                "data": mark_data,
                "cols": cols,
                "indexcol": mark_indexcol,
                "axis": mark.get("axis", "left"),
                "colors": mark.get("colors"),
                "attrs": attrs,
                "xrange": mark.get("xrange"),
                "margins": mark.get("margins"),
                "xformat": mark.get("xformat"),
                "yrange": mark.get("yrange"),
                "ryrange": mark.get("ryrange"),
                "debug": mark.get("debug", spec.get("debug", False)),
            }

            if mark_type == "line":
                chart.lines(
                    linewidth=mark.get("linewidth"),
                    linebreaks=mark.get("linebreaks", False),
                    drawstyle=mark.get("drawstyle"),
                    **common,
                )
            elif mark_type == "bar":
                chart.bars(
                    stack=mark.get("stack", True),
                    total_barwidth=mark.get("total_barwidth"),
                    linewidth=mark.get("linewidth"),
                    edgecolor=mark.get("edgecolor"),
                    **common,
                )
            elif mark_type == "area":
                chart.area(
                    stack=mark.get("stack", True),
                    linewidth=mark.get("linewidth"),
                    edgecolor=mark.get("edgecolor"),
                    alpha=mark.get("alpha", 1),
                    **common,
                )

        for hline in annotations.get("hlines", []) or []:
            chart.hline(**hline)
        for vline in annotations.get("vlines", []) or []:
            chart.vline(**vline)
        for hrect in annotations.get("hrects", []) or []:
            chart.hrect(**hrect)
        for vrect in annotations.get("vrects", []) or []:
            chart.vrect(**vrect)
        for fill in annotations.get("fills", []) or []:
            chart.fill(**fill)
        for text_annotation in annotations.get("texts", []) or []:
            chart.text(**text_annotation)
        for arrow in annotations.get("arrows", []) or []:
            chart.arrow(**arrow)

        chart._spec = spec.copy()
        return chart

    def to_spec(self):
        """
        Return a JSON-like chart specification for the current chart.

        For charts created with from_spec(), this preserves the original spec.
        For charts created imperatively, it reports the stable chart-level
        settings that can be inferred from the object.
        """

        if hasattr(self, "_spec"):
            return self._spec.copy()

        return {
            "data": {"indexcol": self.indexcol},
            "marks": [],
            "text": {
                "title": self.titletext,
                "subtitle": self.subtitletext,
                "xtitle": self.xtitletext,
                "ytitle": self.ytitletext,
            },
            "axes": {
                "x": {
                    "range": getattr(self, "_xrange", None),
                    "format": self.xformat,
                    "margins": self.margins,
                },
                "y": {"range": getattr(self, "_yrange", None)},
                "right_y": {"range": getattr(self, "_ryrange", None)},
            },
            "legend": {
                "show": self.show_legend,
                "ncol": self.ncol_legend,
                "left": self.legend_left,
                "bottom": self.legend_bottom,
                "width": self.legend_width,
                "height": self.legend_height,
                "header": self.legend_header,
            },
            "layout": {
                "width": self.width,
                "height": self.height,
                "topaxis": self._topaxis,
            },
        }

    def __init__(self, data=None, indexcol=None,
                 # plotting options
                 linecols=None, barcols=None, rlinecols=None, areacols=None,
                 # iterable of colors used as cycle
                 colors=None,
                 # lines() options ---------------------------------------------------------
                 # Global linewidth
                 linewidth=None,
                 # whether to remove breaks in line charts
                 linebreaks=False,
                 # Set drawstyle to "steps-post" to draw step plots
                 drawstyle='default',
                 # bars() options ----------------------------------------------------------
                 barstack=True, barwidth=None, baraxis='left',
                 total_barwidth=None,
                 barlinewidth=None,
                 baredgecolor='black',
                 # area() options ----------------------------------------------------------
                 areastack=True, areaaxis='left',
                 arealinewidth=None,
                 areaedgecolor='none',
                 alpha=1,
                 # title options -----------------------------------------------------------
                 title=None,
                 titlecolor=None,
                 titlefontsize=18,
                 titlefont=None,
                 titlefontweight='bold',
                 titleloc='left',
                 titley=1.05,

                 subtitle=None,
                 subtitlecolor='#4B82AD',
                 subtitlefont=None,
                 subtitlefontsize=12,
                 subtitlefontweight='normal',
                 subtitleha='left',
                 subtitleva='center',
                 subtitley=1.03,
                 
                 # h/vlines, h/vrects, texts, arrows, fills --------------------------------
                 hlines=None, vlines=None, hrects=None, vrects=None,
                 fills=None, texts=None, arrows=None,
                 
                 # xtitle, ytitle options --------------------------------------------------
                 xtitle='', 
                 xfontsize=14, xfont='Segoe UI', xfontweight='normal', xcolor='black', xpad=8,
                 xloc='left', xrotation=0, xalpha=1,

                 ytitle='', 
                 yfontsize=14, yfont='Segoe UI', yfontweight='normal', ycolor='black', ypad=8,
                 yloc='center', yrotation=90, yalpha=1,
                 
                 # xtick, ytick options ----------------------------------------------------
                 xtickfontsize=None, ytickfontsize=None,
                 xticklength=None, yticklength=None,
                 xtickpad=None, ytickpad=None,
                 xtickangle=0, ytickangle=0,
                 # Number of ticks on axes.
                 # nticksx is set to 'auto' so that
                 # - for xaxis_type of 'datetime', it will automatically set label positions
                 # - for xaxis_type of 'categorical', it will show all labels
                 # - for xaxis_type of 'numerical' it will defualt to try 7 labels
                 # If an integer is specified, the number of labels is adjusted to roughly match.
                 # 
                 # By default nticksy is None so that we let Matplotlib automatically decide.
                 # Specifying the number of ticks can be useful at times but often leads to
                 # the ticks appearing at numbers that are not as round as the default.
                 nticksx='auto', nticksy=None,
                 # Whether the final period for a time series x-axis is shown as tick and label
                 last=False,
                 # These are for specifying the exact tick locations and labels
                 xticks=None, xlabels=None,
                 yticks=None, ylabels=None,
                 ryticks=None, rylabels=None,
                 
                 # individual look of each column in data-----------------------------------
                 attrs=None,
                 xrange=None, yrange=None, ryrange=None,
                 xformat='auto',
                 margins='auto',
                 width=10, height=6,
                 topaxis='left',
                 
                 # legend options ----------------------------------------------------------
                 legend=True, # Whether to show legend
                 ncol_legend=1,
                 legend_spacing=0.5,
                 legend_fontsize=14,
                 legend_header='', legend_header_color='black', legend_header_fontsize=16,
                 legend_left=0.04, legend_bottom=0.85, legend_width=0.70, legend_height=0.15,
                 legend_mode='expand', legend_frame=False,
                 legend_color='white', legend_alpha=1, legend_shadow=False, legend_fancybox=False,
                 legend_edge='black', legend_linestyle='-', legend_linewidth=1,

                 # watermark options ----------------------------------------------------------
                 watermark=None,
                 wmx=None, wmy=None,
                 wmsize=None, wmfont=None, wmcolor=None, wmalpha=None, wmangle=None, wmzorder=None,

                 # fontname, lang for special languages ----------------------------------------------------------
                 fontname=None,
                 lang=None,

                 debug=False):

        # ------------------------------------------------------------------------------
        # Set attributes
        self.debug = debug
        self.data = data

        # Set self.indexcol
        self.indexcol= indexcol
        if self.indexcol is not None:
            if self.data is not None and type(self.data) == pd.DataFrame:
                if indexcol not in self.data.columns:
                    error = 'indexcol specified as "' + str(indexcol) + '" but not found in data:\n'
                    error += str(data)
                    raise ValueError(error)
                else:
                    if self.debug:
                        print('Setting index to ' + str(indexcol))
                    self.data.set_index(indexcol, inplace=True)
                    # If possible, convert to datetime index
                    try:
                        self.data.index = pd.to_datetime(self.data.index)
                    except Exception:
                        pass
            else:
                print('WARNING: indexcol specified but data is not DataFrame')

        # Set colors using iterable of colors if specified
        if colors is not None:
            self.colors = itertools.cycle([c for c in colors])
        # Otherwise default to style file
        else:
            prop_cycle = [v['color'] for v in plt.rcParams['axes.prop_cycle']]
            self.colors = itertools.cycle(prop_cycle)

        # Set global linewidth, default is None.
        self.linewidth = linewidth

        # Set global line drawstyle
        self.drawstyle = drawstyle
        
        # If barlinewidth was specified, use it
        if barlinewidth is not None:
            self.barlinewidth = barlinewidth
        # Otherwise get from style
        else:
            self.barlinewidth = matplotlib.rcParams['patch.linewidth']
        self.baredgecolor = baredgecolor

        # If arealinewidth was specified, use it
        if arealinewidth is not None:
            self.arealinewidth = arealinewidth
        # Otherwise get from style
        else:
            self.arealinewidth = matplotlib.rcParams['patch.linewidth']
        self.areaedgecolor = areaedgecolor

        self.alpha = alpha
        
        self.width = width
        self.height = height

        # title options
        self.titletext = title # self.title is reserved for function to set title
        self.titlecolor = titlecolor
        self.titlefontsize = titlefontsize
        self.titlefontweight = titlefontweight
        self.titlefont = titlefont
        self.titleloc = titleloc
        self.titley = titley
        
        # subtitle options
        # Keep a handle when a subtitle is added so it can be deleted later
        self.subtitlehandle = None
        self.subtitletext = subtitle # self.subtitle is reserved for function to set title
        self.subtitlecolor = subtitlecolor
        self.subtitlefont = subtitlefont
        self.subtitlefontsize = subtitlefontsize
        self.subtitlefontweight = subtitlefontweight
        self.subtitleha = subtitleha
        self.subtitleva = subtitleva
        self.subtitley = subtitley
        
        # x-axis label options
        self.xtitletext = xtitle
        self.xfontsize = xfontsize
        self.xfont = xfont
        self.xfontweight = xfontweight
        self.xpad = xpad
        self.xcolor = xcolor
        self.xloc = xloc
        self.xrotation = xrotation
        self.xalpha = xalpha

        # y-axis label options
        self.ytitletext = ytitle
        self.yfontsize = yfontsize
        self.yfont = yfont
        self.yfontweight = yfontweight
        self.ypad = ypad
        self.ycolor = ycolor
        self.yloc = yloc
        self.yrotation = yrotation
        self.yalpha = yalpha

        # If params for ticks were specified use them,
        # otherwise default to style
        if xtickfontsize is not None:
            self.xtickfontsize = xtickfontsize
        else:
            self.xtickfontsize = matplotlib.rcParams['xtick.labelsize']

        if ytickfontsize is not None:
            self.ytickfontsize = ytickfontsize
        else:
            self.ytickfontsize = matplotlib.rcParams['ytick.labelsize']

        if xticklength is not None:
            self.xticklength = xticklength
        else:
            self.xticklength = matplotlib.rcParams['xtick.major.size']

        if yticklength is not None:
            self.yticklength = yticklength
        else:
            self.yticklength = matplotlib.rcParams['ytick.major.size']

        if xtickpad is not None:
            self.xtickpad = xtickpad
        else:
            self.xtickpad = matplotlib.rcParams['xtick.major.pad']

        if ytickpad is not None:
            self.ytickpad = ytickpad
        else:
            self.ytickpad = matplotlib.rcParams['ytick.major.pad']

        # No rcParams for angle
        if xtickangle is not None:
            self.xtickangle = xtickangle
        else:
            self.xtickangle = 0

        if ytickangle is not None:
            self.ytickangle = ytickangle
        else:
            self.ytickangle = 0

        # Set nticksx, nticksy
        self.nticksx = nticksx
        self.nticksy = nticksy
        self.last = last

        # Specify tick location, labels
        self.xticks = xticks
        self.xlabels = xlabels
        self.yticks = yticks
        self.ylabels = ylabels
        self.ryticks = ryticks
        self.rylabels = rylabels
        # Check that lengths match
        if xticks is not None or xlabels is not None:
            try:
                len(self.xticks) == len(self.xlabels)
            except Exception as e:
                raise ValueError('xticks and xlabels must be iterables of same length, got exception:' + str(e))
        if yticks is not None or ylabels is not None:
            try:
                len(self.yticks) == len(self.ylabels)
            except Exception as e:
                raise ValueError('yticks and ylabels must be iterables of same length, got exception:' + str(e))
        if ryticks is not None or rylabels is not None:
            try:
                len(self.ryticks) == len(self.rylabels)
            except Exception as e:
                raise ValueError('ryticks and rylabels must be iterables of same length, got exception:' + str(e))
            
        self.xformat = xformat
        self.margins = margins

        # self._legend is None while it does not exist or show_legend=False
        # self.show_legend is the internal variable that controls whether a legned is shown.
        self._legend = None
        self.show_legend = legend
        
        self.ncol_legend = ncol_legend
        self.legend_fontsize = legend_fontsize
        self.legend_header = legend_header
        self.legend_header_color = legend_header_color
        self.legend_header_fontsize = legend_header_fontsize

        # Whether left or right x-axis should be drawn on top.
        # Use self.top_xaxis() to set.
        self._topaxis = topaxis

        # ------------------------------------------------------------------------------
        # Entries, labels for legend
        self.legend_entries = []
        self.legend_labels = []

        self.legend_left = legend_left
        self.legend_bottom = legend_bottom
        self.legend_width = legend_width
        self.legend_height = legend_height
        self.legend_mode = legend_mode
        self.legend_frame = legend_frame
        self.legend_spacing = legend_spacing
        self.legend_frame = False
        self.legend_color = legend_color
        self.legend_alpha = 1
        self.legend_shadow = legend_shadow
        self.legend_fancybox = legend_fancybox
        self.legend_edge = legend_edge
        self.legend_linestyle = legend_linestyle
        self.legend_linewidth = legend_linewidth

        self.fontname = fontname

        # Set language
        self._lang = lang
        self.fontlang(self._lang)
        
        # Set self.xaxis_type based on data
        self.xaxis_type = self._set_xaxis_type()
        
        # Create figure
        self.fig, self.ax = plt.subplots(1, 1, figsize=(self.width, self.height))
        # Initialize self.ax_right to None, only generate as needed
        self.ax_right = None
        if self.debug:
            print('Created self.fig, self.ax')

        # Add title, subtitle
        self.title(self.titletext)
        self.subtitle(self.subtitletext)

        # Add x, y title
        self.xtitle(self.xtitletext)
        self.ytitle(self.ytitletext)

        # Set x, y tick font size
        self.ticks(axis='x')
        self.ticks(axis='y')

        # Get xrange and trim data as needed
        self._xrange = self._parse_xrange(xrange, debug=self.debug)
        self._trim_data(self._xrange, debug=self.debug)

        # Set whether to allow breaks in line charts.
        # If linebreaks is True, NA values will have a break in lines.
        self.linebreaks = linebreaks
        
        # Set x-axis range.
        if self.data is not None:
            self.xrange(self._xrange, self.margins)

        # Set y-axis ranges if specified.
        # If yrange is None, keep as None instead of passing through _parse_yrange()
        # which will set it to (None, None), and when used set the y-axis range to
        # Matplotlib's default (0, 1).
        # Below, self.yrange and self.ryrange are always set,
        # and if None is passed in, they remain None.
        self._yrange = yrange
        if yrange is not None:
            self._yrange = self._parse_yrange(yrange)
            self.yrange(self._yrange)

        self._ryrange = ryrange
        if ryrange is not None:
            self._ryrange = self._parse_yrange(ryrange)
            self.ryrange(self._ryrange)
            
            
        # Set x-axis formatting
        self.xaxis_format()

        # Set which x-axis is drawn on top
        self.topaxis(self._topaxis)

        # List of text annotation handles so text can be modified later
        self.texts = []

        # ---------------------------------------------------------------------------------------------------
        # Draw area, bars, lines
        if areacols is not None:
            self.area(self.data, areacols, indexcol=self.indexcol, axis=areaaxis, colors=None, alpha=self.alpha,
                      stack=areastack, linewidth=self.linewidth, edgecolor=self.areaedgecolor,
                      attrs=attrs,
                      xrange=self._xrange,
                      debug=self.debug)
            
        if barcols is not None:
            self.bars(self.data, barcols, indexcol=self.indexcol, colors=None, stack=barstack, total_barwidth=total_barwidth,
                      axis=baraxis, linewidth=self.barlinewidth, edgecolor=self.baredgecolor,
                      attrs=attrs,
                      xrange=self._xrange,
                      debug=self.debug)
            
        if linecols is not None:
            self.lines(self.data, linecols, indexcol=self.indexcol, colors=None,
                       linewidth=self.linewidth, linebreaks=self.linebreaks, drawstyle=self.drawstyle,
                       attrs=attrs,
                       xrange=self._xrange,
                       debug=self.debug)

        if rlinecols is not None:
            self.lines(self.data, rlinecols, indexcol=self.indexcol, axis='right', colors=None,
                       linewidth=self.linewidth, linebreaks=self.linebreaks, drawstyle=self.drawstyle,
                       attrs=attrs, 
                       xrange=self._xrange,
                       debug=self.debug)
            
        # If hlines, vlines, hrects, vrects, fills, texts, arrows is given, loop over.
        # Each of these should be a list of kwargs of the form
        # [{'y' : 10, 'color' : 'red'}, ...]
        if hlines is not None:
            try:
                for hline in hlines:
                    self.hline(**hline)
            except Exception as e:
                raise RuntimeError('Could not process hlines = ' + str(hlines) + ' with exception:' + str(e))

        if vlines is not None:
            try:
                for vline in vlines:
                    self.vline(**vline)
            except Exception as e:
                raise RuntimeError('Could not process vlines = ' + str(vlines) + ' with exception:' + str(e))

        if hrects is not None:
            try:
                for hrect in hrects:
                    self.hrect(**hrect)
            except Exception as e:
                raise RuntimeError('Could not process hrects = ' + str(hrects) + ' with exception:' + str(e))

        if vrects is not None:
            try:
                for vrect in vrects:
                    self.vrect(**vrect)
            except Exception as e:
                raise RuntimeError('Could not process vrects = ' + str(vrects) + ' with exception:' + str(e))

        if fills is not None:
            try:
                for fill in fills:
                    self.fill(**fill)
            except Exception as e:
                raise RuntimeError('Could not process fills = ' + str(fills) + ' with exception:' + str(e))
            
        if texts is not None:
            try:
                for text in texts:
                    self.text(**text)
            except Exception as e:
                raise RuntimeError('Could not process texts = ' + str(texts) + ' with exception:' + str(e))

        if arrows is not None:
            try:
                for arrow in arrows:
                    self.arrow(**arrow)
            except Exception as e:
                raise RuntimeError('Could not process arrows = ' + str(arrows) + ' with exception:' + str(e))

        # Create legend
        if self.debug:
            print('self.legend_entries:')
            print(self.legend_entries)
            print('self.legend_labels:')
            print(self.legend_labels)
        if self.show_legend:
            self.legend()

        # Add watermark
        if watermark is not None:
            text = watermark
            self.watermark(text=text,x=wmx, y=wmy,
                           size=wmsize, font=wmfont, color=wmcolor, alpha=wmalpha, angle=wmangle, zorder=wmzorder)

    def _repr_png_(self):
        return self.fig._repr_png_()

    def _ipython_display_(self):
        display(self.fig)
    
    def apply(self, style):
        '''
        Apply style to this Chart.
        '''
        
        pass

    def _set_xaxis_type(self):

        # Check whether index of data is datetime
        if self.data is not None and type(self.data) == pd.DataFrame:
            # Try to convert to datetime axis
            try:
                # Ignore pandas warnings
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    self.data.index = pd.to_datetime(self.data.index)
                xaxis_type = 'datetime'
            except:
                pass
            
            idx_types = {type(idx) for idx in self.data.index}
            if idx_types in [{pd.Timestamp}, {pd.Period}]:
                xaxis_type = 'datetime'
            elif idx_types in [{int}, {float}]: 
                xaxis_type = 'numerical'
            else:
                xaxis_type = 'categorical'
        else:
            xaxis_type = None

        self.xaxis_type = xaxis_type
            
        # Return so it is clear what was set
        return xaxis_type

    def _parse_xrange(self, xrange, debug=False):
        '''
        Utility function to parse x-axis range.
        This can either be a date range or a number range.
        Returns a list of two elements, the elements are one of
        - a number
        - a str interpretible as a date
        - None
        '''

        if debug:
            print('start of _parse_xrange with xrange:')
            print(xrange)

        if xrange is None:
            if debug:
                print('returning [None, None]')
            return [None, None]

        # Special case of two floats or ints, just return original.
        try:
            x0, x1 = xrange
            if type(x0) in [float, int] and type(x1) in [float, int]:
                if debug:
                    print('returning:')
                    print(xrange)
                return xrange
        except Exception:
            pass
        
        if type(xrange) == str and xrange.find(':') != -1:
            xrange = [x.strip() for x in xrange.split(':', 1)]
            xrange = [x if x != '' else None for x in xrange]
            if debug:
                print('xrange here:')
                print(xrange)

        if self.xaxis_type == 'datetime':
            try:
                xrange = [pd.Timestamp(str(x)) if x is not None else None for x in xrange]
                if debug:
                    print('returning:')
                    print(xrange)
            except ValueError:
                raise ValueError('Could not convert xrange = ' + str(xrange) + ' to pd.Timestamp')
            
        return [xrange[0], xrange[1]]

    def _parse_yrange(self, yrange):
        '''
        Utility function to parse y-axis range.
        This must be a number range.
        
        Input can be one of
        - None        Return None, no calculation done.
        - (y1, y2)    Returns provided values y1, y2.
                      If either is None, this side will default to available value in df.
        '''

        if yrange is None:
            return [None, None]

        return [yrange[0], yrange[1]]

    def _trim_data(self, xrange, debug=False):
        '''
        Trim self.data using xrange.
        '''

        if debug:
            print('start of _trim_data() for data:')
            print(self.data)
            print('xrange = ' + str(xrange))
        
        
        if type(self.data) == pd.DataFrame:
            if xrange[0] is None and xrange[1] is None:
                # Don't trim if nothing specified
                pass
            elif xrange[0] is None:
                # Trim start only
                self.data = self.data.loc[:str(xrange[1]), :]
            elif xrange[1] is None:
                # Trim end only
                self.data = self.data.loc[str(xrange[0]):, :]
            else:
                # Trim both
                self.data = self.data.loc[str(xrange[0]):str(xrange[1]), :]

            if debug:
                print('After applying trim, self.data:')
                print(self.data)

    def title(self, text, color=None, fontsize=None, font=None, fontweight=None,
              loc=None, y=None,
              debug=False):

        # Use text if it is given, otherwise use self.title.
        # If both are None don't do anything
        if text is None:
            text = self.titletext
            if text is None:
                return

        # Save as attribute
        self.ttitletext = text

        # For other params use rcParams, if self attributes are set, override.
        # Further override with input params.
        if fontsize is None:
            fontsize = self.titlefontsize
            if fontsize is None:
                fontsize = matplotlib.rcParams['axes.titlesize']

        if color is None:
            color = self.titlecolor
            if color is None:
                color = matplotlib.rcParams['axes.titlecolor']
                # Check whether color is 'auto', if so fall back
                # on text.color
                if color == 'auto':
                    color = matplotlib.rcParams['text.color']
                    # Fall back further
                    if color == 'auto':
                        color = 'black'
                
        if font is None:
            font = self.titlefont
            if font is None:
                if self.fontname is not None:
                    font = self.fontname

        if fontweight is None:
            fontweight = self.titlefontweight
            if fontweight is None:
                fontweight = matplotlib.rcParams['figure.titleweight']

        if loc is None:
            loc = self.titleloc
            if loc is None:
                # No rcParam, default to left
                loc = 'left'

        if y is None:
            y = self.titley
            if y is None:
                y = matplotlib.rcParams['axes.titley']


        if debug:
            print('title = ' + str(text) + ' color = ' + str(color) + ' loc = ' + str(loc) + ' y = ' + str(y))
            print('fontweight = ' + str(fontweight) + ' fontname = ' + str(font) + ' fontsize = ' + str(fontsize))

        if font is not None:
            self.ax.set_title(str(text), color=color, loc=loc, y=y,
                              fontweight=fontweight, fontname=font, fontsize=fontsize)
        else:
            self.ax.set_title(str(text), color=color, loc=loc, y=y,
                              fontweight=fontweight, fontsize=fontsize)

    def subtitle(self, text, color=None, fontsize=None, font=None, fontweight=None,
                 ha=None, va=None, y=None,
                 debug=False):

        # Use text if it is given, otherwise use self.subtitle.
        # If both are None don't do anything
        if text is None:
            text = self.subtitletext
            if text is None:
                return

        # Save as attribute
        self.subttitletext = text
            
        # For other params use rcParams, if self attributes are set, override.
        # Further override with input params.
        if fontsize is None:
            fontsize = self.subtitlefontsize
            if fontsize is None:
                fontsize = 12

        if color is None:
            color = self.subtitlecolor
            if color is None:
                color = matplotlib.rcParams['axes.titlecolor']
                
        if font is None:
            font = self.subtitlefont
            if font is None:
                if self.fontname is not None:
                    font = self.fontname

        if fontweight is None:
            fontweight = self.subtitlefontweight
            if fontweight is None:
                fontweight = matplotlib.rcParams['figure.titleweight']

        if ha is None:
            ha = self.subtitleha
            if ha is None:
                # No rcParam, default to left
                ha = 'left'

        if va is None:
            va = self.subtitleva
            if va is None:
                # No rcParam, default to center
                va = 'center'
                
        if y is None:
            y = self.subtitley
            if y is None:
                y = 1.03

        if debug:
            print('text = ' + str(text) + ' color = ' + str(color) + ' ha = ' + str(ha) + ' va = ' + str(va) + ' y = ' + str(y))
            print('fontweight = ' + str(fontweight) + ' fontname = ' + str(font) + ' fontsize = ' + str(fontsize))

        if font is not None:
            self.subtitlehandle = self.ax.text(0., y, text,
                                               color=color, fontsize=fontsize, fontweight=fontweight, fontname=font,
                                               horizontalalignment=ha, verticalalignment=va, transform=self.ax.transAxes)
        else:
            self.subtitlehandle = self.ax.text(0., y, text,
                                               color=color, fontsize=fontsize, fontweight=fontweight,
                                               horizontalalignment=ha, verticalalignment=va, transform=self.ax.transAxes)
        
    def xtitle(self, text, fontsize=None, font=None, fontweight=None, color=None, pad=None, loc=None,
               rotation=None, alpha=1):

        # Get text to show
        if text is None:
            # Use internal xtitletext
            text = self.xtitletext

            # If nothing specified, quit
            if text is None:
                return
        else:
            self.xtitletext = text

        # For all other params, hierarchy is
        # 1. specified within function
        # 2. class attributes set in __init__()
        # 3. rcParams

        # fontsize
        if fontsize is None:
            fontsize = self.xfontsize
            if fontsize is None:
                fontsize = matplotlib.rcParams['axes.labelsize']
        self.xfontsize = fontsize

        # font
        if font is None:
            font = self.xfont
            if font is None:
                font = matplotlib.rcParams['font.family']
        self.xfont = font

        # fontweight
        if fontweight is None:
            fontweight = self.xfontweight
            if fontweight is None:
                fontweight = matplotlib.rcParams['axes.labelweight']
        self.xfontweight = fontweight

        # color
        if color is None:
            color = self.xcolor
            if color is None:
                color = matplotlib.rcParams['axes.labelcolor']
        self.xcolor = color

        # pad
        if pad is None:
            pad = self.xpad
            if pad is None:
                pad = matplotlib.rcParams['axes.labelpad']
        self.xpad = pad

        # loc
        if loc is None:
            loc = self.xloc
            if loc is None:
                loc = 'left' # no rcParams
        self.xloc = loc

        # rotation
        if rotation is None:
            rotation = self.xrotation
            if rotation is None:
                rotation = 0 # no rcParams
        self.xrotation = rotation

        # alpha
        if alpha is None:
            alpha = self.xalpha
            if alpha is None:
                alpha = 1 # no rcParams
        self.xalpha = alpha
        
        # Set x-axis title
        if self.fontname is not None:
            self.ax.set_xlabel(text, fontsize=fontsize, font=font, fontweight=fontweight,
                               color=color, labelpad=pad, loc=loc, rotation=rotation, alpha=alpha,
                               fontname=self.fontname)
        else:
            self.ax.set_xlabel(text, fontsize=fontsize, fontweight=fontweight,
                               color=color, labelpad=pad, loc=loc, rotation=rotation, alpha=alpha)
        
    def ytitle(self, text, fontsize=None, font=None, fontweight=None, color=None, pad=None, loc=None,
               rotation=None, alpha=1):

        # Get text to show
        if text is None:
            # Use internal ytitletext
            text = self.ytitletext

            # If nothing specified, quit
            if text is None:
                return
        else:
            self.ytitletext = text

        # For all other params, hierarchy is
        # 1. specified within function
        # 2. class attributes set in __init__()
        # 3. rcParams

        # fontsize
        if fontsize is None:
            fontsize = self.yfontsize
            if fontsize is None:
                fontsize = matplotlib.rcParams['axes.labelsize']
        self.yfontsize = fontsize

        # font
        if font is None:
            font = self.yfont
            if font is None:
                font = matplotlib.rcParams['font.family']
        self.yfont = font

        # fontweight
        if fontweight is None:
            fontweight = self.yfontweight
            if fontweight is None:
                fontweight = matplotlib.rcParams['axes.labelweight']
        self.yfontweight = fontweight

        # color
        if color is None:
            color = self.ycolor
            if color is None:
                color = matplotlib.rcParams['axes.labelcolor']
        self.ycolor = color

        # pad
        if pad is None:
            pad = self.ypad
            if pad is None:
                pad = matplotlib.rcParams['axes.labelpad']
        self.ypad = pad

        # loc
        if loc is None:
            loc = self.yloc
            if loc is None:
                loc = 'center' # no rcParams
        self.yloc = loc

        # rotation
        if rotation is None:
            rotation = self.yrotation
            if rotation is None:
                rotation = 90 # no rcParams
        self.yrotation = rotation

        # alpha
        if alpha is None:
            alpha = self.yalpha
            if alpha is None:
                alpha = 1 # no rcParams
        self.yalpha = alpha
        
        # Set y-axis title
        if self.fontname is not None:
            self.ax.set_ylabel(text, fontsize=fontsize, font=font, fontweight=fontweight,
                               color=color, labelpad=pad, loc=loc, rotation=rotation, alpha=alpha,
                               fontname=self.fontname)
        else:
            self.ax.set_ylabel(text, fontsize=fontsize, fontweight=fontweight,
                               color=color, labelpad=pad, loc=loc, rotation=rotation, alpha=alpha)

    def ticks(self, axis='x', yaxis='left', size=None, length=None, angle=None, pad=None,
              nticks=None, last=None,
              ticks=None, labels=None,
              debug=False):
        '''
        Control ticks. Pass in axis of "x" or "y" (for right y-axis pass in yaxis='right'),
        together with
        - size   : tick font size
        - length : tick length
        - angle  : tick font angle
        - pad    : padding between ticks and tick labels
        --------------
        Below are options for specifying custom tick positions and labels.
        If using option `last` does not provide good results, it may be necessary
        to specify custom ticks and labels.
        If ticks and labels are specified, inputs for nticks and last are ignored.
        - nticks : number of ticks along axis. If "auto", will depend of xaxis_type
        - last   : If True, ensure last x-axis tick aligns with final data point
        - ticks  : Specify an iterable of tick positions.
        - labels : Specify an iterable of tick labels. Length must match ticks.
        '''

        # ----------------------------------------------------------------------------------------------------------------
        # Check input options
        if axis not in ['x', 'y']:
            raise ValueError('axis must be "x" or "y", given ' + str(axis))

        if yaxis not in ['left', 'right']:
            raise ValueError('yaxis must be "left" or "right", given ' + str(yaxis))

        if axis == 'x' and yaxis == 'right':
            raise ValueError('axis given as "x" and yaxis given as "right"')

        # When formatting y-axis, ensure locale is used.
        # This will be set from the lang() function in __init__.py
        self.ax.yaxis.set_major_formatter(ScalarFormatter(useLocale=True))
        
        if nticks is None:
            if axis == 'x':
                nticks = self.nticksx
            else:
                nticks = self.nticksy
        else:
            if axis == 'x':
                self.nticksx = nticks
            else:
                self.nticksy = nticks

        if last is None:
            last = self.last

        if ticks is None:
            if axis == 'x':
                ticks = self.xticks
            else:
                if yaxis == 'left':
                    ticks = self.yticks
                else:
                    ticks = self.ryticks
        else:
            if axis == 'x':
                self.xticks = ticks
            else:
                if yaxis == 'left':
                    self.yticks = ticks
                else:
                    self.ryticks = ticks
            
        if labels is None:
            if axis == 'x':
                labels = self.xlabels
            else:
                if yaxis == 'left':
                    labels = self.ylabels
                else:
                    labels = self.rylabels
        else:
            if axis == 'x':
                self.xlabels = labels
            else:
                if yaxis == 'left':
                    self.ylabels = labels
                else:
                    self.rylabels = labels

        # If ticks or labels are specified, make sure they have same length
        if ticks is not None or labels is not None:
            try:
                len(ticks) == len(labels)
            except Exception:
                raise ValueError('ticks and labels must be iterables with same length')
        
        # ----------------------------------------------------------------------------------------------------------------
        # Set tick length, angle, padding, font size
        # Options used are size, length, pad, angle.
        
        # If params are specified, update internal value,
        # otherwise use class values.

        # tick label size
        if size is not None:
            if axis == 'x':
                self.xtickfontsize = size
            else:
                self.ytickfontsize = size
        else:
            if axis == 'x':
                size = self.xtickfontsize
            else:
                size = self.ytickfontsize
            
        # tick length
        if length is not None:
            if axis == 'x':
                self.xticklength = length
            else:
                self.yticklength = length
        else:
            if axis == 'x':
                length = self.xticklength
            else:
                length = self.yticklength

        # padding between tick and labels
        if pad is not None:
            if axis == 'x':
                self.xtickpad = pad
            else:
                self.ytickpad = pad
        else:
            if axis == 'x':
                pad = self.xtickpad
            else:
                pad = self.ytickpad

        # angle of tick labels
        if angle is not None:
            if axis == 'x':
                self.xtickangle = angle
            else:
                self.ytickangle = angle
        else:
            if axis == 'x':
                angle = self.xtickangle
            else:
                angle = self.ytickangle

        if debug:
            print('length = ' + str(length) + ' size = ' + str(size) + ' pad = ' + str(pad) + ' angle = ' + str(angle))
                
        # Set
        if yaxis == 'left':
            self.ax.tick_params(axis=axis, which='major',
                                length=length, labelsize=size,
                                pad=pad, labelrotation=angle)
        else:
            if self.ax_right:
                self.ax_right.tick_params(axis=axis, which='major',
                                length=length, labelsize=size,
                                pad=pad, labelrotation=angle)
            else:
                print('WARNING: yaxis="right" was specified but no right y-axis exists')

        # ----------------------------------------------------------------------------------------------------------------
        # Set tick positions and labels.
        # Options used are nticks, last, ticklist, labellist.

        # If ticklist and labellist are specified, use these and do not use last or nticks.
        if ticks is not None:
            if axis == 'x':
                # If datetime x-axis, try to convert specified values to dates
                if self.xaxis_type == 'datetime':
                    try:
                        tick_dates = [pd.Timestamp(x) for x in ticks]
                    except Exception:
                        tick_dates = ticks
                self.ax.set_xticks(tick_dates)
                self.ax.set_xticklabels(labels)
            elif axis == 'y':
                if yaxis == 'left':
                    self.ax.set_yticks(ticks)
                    self.ax.set_yticklabels(labels)
                else:
                    if self.ax_right:
                        self.ax_right.set_yticks(ticks)
                        self.ax_right.set_yticklabels(labels)
                    else:
                        print('WARNING: ticks and labels specified for right y-axis but does not exist')
                        
            # If ticks or labels were specified, ignore nticks and last
            return

        # If no ticks or labels were specified, use nticks and last
        if axis == 'x':
            if nticks is not None:
                self.nticksx = nticks

            # Integer
            if type(self.nticksx) == int:
                # This is not implemented as specifying MaxNLocator(nbins)
                # will not align the dates with period start dates as with other charts.
                # The proper way to do this would be to add this option into _set_datetime_ticks()
                # with the appropriate argument.
                # self.ax.xaxis.set_major_locator(MaxNLocator(nbins=self.nticksx))
                raise ValueError('Cannont specify int value for nticks for datetime x-axis')
            elif self.nticksx == 'auto':
                # Check xaxis_type
                if self.xaxis_type == 'datetime':
                    self._set_datetime_ticks(debug=debug)
                elif self.xaxis_type == 'categorical':
                    if self.data is not None:
                        self.ax.set_xticks(np.arange(len(self.data)), labels=self.data.index)
                elif self.xaxis_type == 'numerical':
                    self.ax.xaxis.set_major_locator(MaxNLocator(nbins=7))
                elif self.xaxis_type is None:
                    # In some cases for example if no data was passed in,
                    # no need to set.
                    pass
                else:
                    raise ValueError('self.xaxis_type was ' + str(self.xaxis_type))
            else:
                raise ValueError('nticks for x-axis must be int or "auto", given ' + str(nticks))

            # If last was specified for a datetime axis, replace last tick with final data point.
            # This ensures that number of ticks follows standard.
            if self.xaxis_type == 'datetime' and last:
                # Get current ticks
                ticks = [pd.Timestamp(x.replace(tzinfo=None)) for x in mdates.num2date(self.ax.get_xticks())]
                if debug:
                    print('original ticks:')
                    print(ticks)
                # Replace final eleemnt from data
                x = self.data.index[-1]
                ticks = ticks[:-1] + [x]
                self.ax.set_xticks(ticks)
                if debug:
                    print('new ticks:')
                    print(ticks)
        # end of axis == 'x'
        else:
            if nticks is not None:
                self.nticksy = nticks

            if self.nticksy is not None:
                if yaxis == 'left':
                    self.ax.yaxis.set_major_locator(MaxNLocator(nbins=self.nticksy))
                else:
                    if self.ax_right:
                        self.ax_right.set_major_locator(MaxNLocator(nbins=self.nticksy))
                    else:
                        print('WARNING: yaxis="right" was specified but no right y-axis exists')
                
    def _set_datetime_ticks(self, debug=False):
        '''
        For datetime x-axis, set location of ticks based on freq and length of data.
        This is called from ticks() when self.axis_type is "datetime" and self.nticks is "auto".
        '''
        
        # From freq try to specify location of ticks
        freq = guess_freq(self.data)
        if debug:
            print('freq = ' + str(freq) + ', len(self.data) = ' + str(len(self.data)))
            
        if freq == 'D':
            if len(self.data) <= 30:
                self.ax.xaxis.set_major_locator(mdates.DayLocator([1, 8, 15, 22, 29]))
            elif 30 < len(self.data) and len(self.data) <= 90:
                self.ax.xaxis.set_major_locator(mdates.DayLocator([1, 15]))
            elif len(self.data) <= 180:
                self.ax.xaxis.set_major_locator(mdates.DayLocator([1]))
            elif len(self.data) <= 360:
                self.ax.xaxis.set_major_locator(mdates.MonthLocator([1, 3, 5, 7, 9, 11]))
            elif len(self.data) <= 720:
                self.ax.xaxis.set_major_locator(mdates.MonthLocator([1, 4, 7, 10]))
            elif len(self.data) <= 1260:
                self.ax.xaxis.set_major_locator(mdates.MonthLocator([1, 7]))
            else:
                # Show only each Jan, but put in interval
                nyears = len(self.data) // 365 + 1
                interval =  nyears // 7
                self.ax.xaxis.set_major_locator(mdates.MonthLocator(bymonth=[1], interval=interval))
        elif freq == 'W':
            if len(self.data) <= 7:
                self.ax.xaxis.set_major_locator(mdates.DayLocator([1, 8, 15, 22, 29]))
            elif len(self.data) <= 15:
                self.ax.xaxis.set_major_locator(mdates.DayLocator([1, 15]))
            elif len(self.data) <= 30:
                self.ax.xaxis.set_major_locator(mdates.MonthLocator([1]))
            elif len(self.data) <= 60:
                self.ax.xaxis.set_major_locator(mdates.MonthLocator([1, 3, 5, 7, 9, 11]))
            elif  len(self.data) <= 80:
                self.ax.xaxis.set_major_locator(mdates.MonthLocator([1, 4, 7, 10]))
            elif len(self.data) <= 180:
                self.ax.xaxis.set_major_locator(mdates.MonthLocator([1, 7]))
            elif len(self.data) <= 360:
                self.ax.xaxis.set_major_locator(mdates.MonthLocator([1]))
            else:
                # Show only each Jan, but put in interval
                nyears = len(self.data) * 7 // 365 + 1
                interval = nyears // 7
                self.ax.xaxis.set_major_locator(mdates.MonthLocator(bymonth=[1], interval=interval))
        elif freq == 'M':
            if len(self.data) <= 7:
                self.ax.xaxis.set_major_locator(mdates.MonthLocator())
            elif len(self.data) <= 12:
                self.ax.xaxis.set_major_locator(mdates.MonthLocator([1, 3, 5, 7, 9, 11]))
            elif len(self.data) <= 24:
                self.ax.xaxis.set_major_locator(mdates.MonthLocator([1, 4, 7, 10]))
            elif len(self.data) <= 48:
                self.ax.xaxis.set_major_locator(mdates.MonthLocator([1, 7]))
            elif len(self.data) < 80:
                self.ax.xaxis.set_major_locator(mdates.MonthLocator([1]))
            elif len(self.data) < 160:
                self.ax.xaxis.set_major_locator(mdates.YearLocator(base=2))
            else:
                nyears = len(self.data) * 30 // 365 + 1
                interval = nyears // 7
                self.ax.xaxis.set_major_locator(mdates.YearLocator(base=interval))
                formatter = mdates.DateFormatter('%b-%y')
                self.ax.xaxis.set_major_formatter(formatter)
        elif freq == 'Q':
            if len(self.data) < 6:
                self.ax.xaxis.set_major_locator(mdates.MonthLocator([1, 4, 7, 10]))
            elif len(self.data) < 12:
                self.ax.xaxis.set_major_locator(mdates.MonthLocator([1, 7]))
            elif len(self.data) < 30:
                self.ax.xaxis.set_major_locator(mdates.MonthLocator([1]))
            elif len(self.data) < 70:
                self.ax.xaxis.set_major_locator(mdates.YearLocator(base=2))
            else:
                nyears = len(self.data) * 90 // 365 + 1
                interval = nyears // 7
                self.ax.xaxis.set_major_locator(mdates.YearLocator(base=interval))
            def quarter_formatter(x, pos):
                date = mdates.num2date(x)
                quarter = (date.month - 1) // 3 + 1
                return f"{date.year}Q{quarter}"
            formatter = FuncFormatter(quarter_formatter)
            self.ax.xaxis.set_major_formatter(formatter)
        elif freq == 'A':
            if len(self.data) < 8:
                self.ax.xaxis.set_major_locator(mdates.YearLocator(base=1))
            elif len(self.data) < 20:
                self.ax.xaxis.set_major_locator(mdates.YearLocator(base=2))
            else:
                base = int(len(self.data) / 6)
                self.ax.xaxis.set_major_locator(mdates.YearLocator(base=base))
            
    def xrange(self, xrange, margins='auto', debug=False):
        '''
        Set x-axis range of fig.
        '''

        if debug:
            print('start of xrange() for xrange = ' + str(xrange) + ', margins = ' + str(margins))
        xrange = self._parse_xrange(xrange)
        if debug:
            print('xrange after calling _parse_xrange():')
            print(xrange)

        # If margins is "auto", try to guess from freq of data
        if margins == 'auto':
            freq = guess_freq(self.data)
            if freq == 'D':
                margins = 1
            elif freq == 'W':
                margins = 5
            elif freq == 'M':
                margins = 30
            elif freq == 'Q':
                margins = 85
            elif freq == 'A':
                margins = 335
            elif freq == '?':
                pass
            else:
                raise ValueError('Unknown freq ' + str(freq))
        else:
            if type(margins) == int:
                # Use input int as-is
                pass
            else:
                raise ValueError('xrange: margins of type ' + str(type(margins)) + ' not allowed')

        if self.xaxis_type == 'datetime':
            if xrange[0] is None:
                xrange[0] = self.data.index.min()
            if xrange[1] is None:
                xrange[1] = self.data.index.max()
        elif self.xaxis_type == 'categorical':
            if xrange[0] is None:
                xrange[0] = 0
            if xrange[1] is None:
                xrange[1] = len(self.data)
        elif self.xaxis_type == 'numerical':
            if xrange[0] is None:
                xrange[0] = self.data.index.min()
            if xrange[1] is None:
                xrange[1] = self.data.index.max()
        elif self.xaxis_type is None:
            return None
        else:
            raise ValueError('self.xaxis_type of ' + str(self.axis_type) + ' not implemented for xrange()')
            
        if debug:
            print('xrange before setting xlim:')
            print(xrange)

        # Trim the data based on xrange
        if self.xaxis_type == 'datetime':
            self._trim_data(xrange)
            
            # Try to interpret xrange as dates and add margin
            try:
                x1, x2 = pd.Timestamp(xrange[0]) - pd.Timedelta(days=margins), pd.Timestamp(xrange[1]) + pd.Timedelta(days=margins)
                self.ax.set_xlim(x1, x2)
                if debug:
                    print('Set xlim to Timestamps')
            # If fails, use xrange as-is.
            except ValueError:
                self.ax.set_xlim(xrange[0],xrange[1])
                if debug:
                    print('Set xlim to xrange as-is')
        elif self.xaxis_type == 'categorical':
            # Default is to add 0.5 margins unless a margin is specified as a float
            try:
                margins = float(margins)
            except:
                margins = 0
            # The second minus sign is not a mistake, need to shift both ranges to the left.
            # The margins are then added to both sides equally.
            self.ax.set_xlim(xrange[0] - 0.5 - margins, xrange[1] - 0.5 + margins)
        elif self.xaxis_type == 'numerical':
            # Default is to add 3% margins unless a margin is specified as a float
            total = np.abs(xrange[1] - xrange[0])
            try:
                margins = float(margins)
                self.ax.set_xlim(xrange[0] - margins * total, xrange[1] + margins * total)
            except:
                self.ax.set_xlim(xrange[0] - 0.03 * total, xrange[1] + 0.03 * total)
        else:
            raise ValueError('self.axis_type of ' + str(self.axis_type) + ' not implemented for xrange()')

    def yrange(self, yrange):
        '''
        Set y-axis range of fig.
        '''

        # Get values from self.ax
        ymin, ymax = self.ax.get_ylim()

        if yrange is None:
            return

        # Set ymin, ymax if they are specified
        if yrange[0] is not None:
            ymin = yrange[0]
            
        if yrange[1] is not None:
            ymax = yrange[1]

        try:
            self.ax.set_ylim(ymin, ymax)
        except Exception as e:
            raise ValueError('Could not apply yrange() for yrange = ' + str(yrange))

    def ryrange(self, ryrange):
        '''
        Set y-axis range of fig.
        '''

        if self.ax_right is None:
            return

        # Get values from self.ax_right
        ymin, ymax = self.ax_right.get_ylim()

        if ryrange is None:
            return

        # Set ymin, ymax if they are specified
        if ryrange[0] is not None:
            ymin = ryrange[0]
            
        if ryrange[1] is not None:
            ymax = ryrange[1]

        try:
            self.ax_right.set_ylim(ymin, ymax)
        except Exception as e:
            raise ValueError('Could not apply ryrange() for yrange = ' + str(yrange))

    def xaxis_format(self, xformat=None, debug=False):
        '''
        Set formatting for datetime x-axis.
        `formatter` should be something like a
        mdates.DateFormatter().

        if xformat is specified, uses it to set formatting.
        If default `auto` is chosen, freq of data is used.
        Specify special options "D", "W", "M", "Q", "A" to set to these frequencies.
        '''

        if debug:
            print('Start of xaxis_format()')

        # If xformat is given use it, otherwise use self.xformat
        if xformat is not None:
            pass
        else:
            xformat = self.xformat

        if debug:
            print('xformat = ' + str(xformat))

        # If self.xaxis_type is not "datetime", skip
        if self.xaxis_type != 'datetime':
            if debug:
                print('self.xaxis_type = ' + str(self.xaxis_type) + ', returning')
            return

        # If default, guess from freq of data.
        if xformat == 'auto':
            # If no data is set, don't try to set
            if self.data is None:
                return
            
             # If length of data is less than 2, cannot calculate difference,
            # use date
            if type(self.data) == pd.DataFrame:
                freq = guess_freq(self.data)
                if freq == 'D':
                    formatter = mdates.DateFormatter('%m/%d/%Y')
                elif freq == 'W':
                    formatter = mdates.DateFormatter('%m/%d/%Y')
                elif freq == 'M':
                    formatter = mdates.DateFormatter('%b-%y')
                elif freq == 'Q':
                    def quarter_formatter(x, pos):
                        date = mdates.num2date(x)
                        quarter = (date.month - 1) // 3 + 1
                        return f"{date.year}Q{quarter}"
                    formatter = FuncFormatter(quarter_formatter)
                    # Below does not work
                    # formatter = mdates.DateFormatter('%YQ%q')
                # Annual
                elif freq == 'A':
                    formatter = mdates.DateFormatter('%Y')
                elif freq == '?':
                    pass
            elif self.data is None:
                # If no data is set, format of x-axis is not set.
                formatter = None
            else:
                raise ValueError('Cannot determine formatter for xformat = auto')
        elif xformat in ['D', 'W']:
            # IMF default for daily
            formatter = mdates.DateFormatter('%m/%d/%Y')
            if debug:
                print('formatter for D:')
                print(formatter)
        elif xformat == 'M':
            # IMF default for monthly
            formatter = mdates.DateFormatter('%b-%y')
            if debug:
                print('formatter for M:')
                print(formatter)
        elif xformat == 'Q':
            # IMF default for quarterly
            def quarter_formatter(x, pos):
                date = mdates.num2date(x)
                quarter = (date.month - 1) // 3 + 1
                return f"{date.year}Q{quarter}"
            formatter = mdates.FuncFormatter(quarter_formatter)
            # Below does not work
            # formatter = mdates.DateFormatter('%YQ%q')
            if debug:
                print('formatter for Q:')
                print(formatter)
        elif xformat == 'A':
            formatter = mdates.DateFormatter('%Y')
            if debug:
                print('formatter for A:')
                print(formatter)
        # Otherwise if specified
        elif type(xformat) == str:
            try:
                formatter = mdates.DateFormatter(xformat)
                if debug:
                    print('formatter for str:')
                    print(formatter)
            except Exception as e:
                raise ValueError('Cannot specify mdates.DateFormatter with xformat ="' + xformat + '"')
        else:
            error = 'xaxis_format():\n'
            error += 'xformat of "' + str(xformat) + '" of type ' + str(type(xformat)) + ' not allowed'
            raise ValueError(error)

        if formatter is not None:
            self.ax.xaxis.set_major_formatter(formatter)

    def lines(self, data=None, cols=None, indexcol=None, axis='left', colors=None, linewidth=None, linebreaks=False, drawstyle=None,
              attrs=None,
              xrange=None, margins=None, xformat=None, yrange=None, ryrange=None,
              debug=False):
        '''
        Add line to chart
        '''

        if debug:
            print('Calling add_lines on "' + str(cols) + '"')

        # Set data.
        # If not specified use self.data.
        if data is None:
            data = self.data

        if indexcol is not None:
            if type(data) == pd.DataFrame:
                # Check if index has already been set for data
                if data.index.name == indexcol:
                    pass

                # Otherwise try to set index to indexcol
                elif indexcol not in data.columns:
                    error = 'indexcol specified as "' + str(indexcol) + '" but not found in data:\n'
                    error += str(data)
                    raise ValueError(error)
                else:
                    data.set_index(indexcol, inplace=True)
                    # If possible, convert to datetime index
                    try:
                        data.index = pd.to_datetime(data.index)
                    except Exception:
                        pass
            else:
                print('WARNING: indexcol specified but data is not DataFrame')
            
        linecols = _parse_cols(cols)
        # Don't try to plot indexcol if it was given since this is now the index
        if indexcol is not None:
            if indexcol in linecols:
                linecols.remove(indexcol)
        if debug:
            print('linecols:')
            print(linecols)

        # If no data and/or linecols are available, finish.
        if data is None:
            if linecols == []:
                print('WARNING: no data given to lines()')
            else:
                print('WARNING: no data given to lines() but linecols given as ' + str(linecols))
            print('No lines will be added')
            return

        # Set self.data to be input data
        self.data = data
        # Set self.xaxis_type from self.data
        self.xaxis_type = self._set_xaxis_type()
        if debug:
            print('self.xaxis_type = ' + str(self.xaxis_type))
            
        # Set x-axis limits if specified, or use default
        if xrange is not None:
            xrange = self._parse_xrange(xrange, debug=debug)
            self._xrange = xrange
        else:
            xrange = self._xrange
        if debug:
            print('xrange:')
            print(xrange)
            
        # Trim data as needed            
        self._trim_data(xrange, debug=debug)
        if debug:
            print(self.data)

        # Create cycle if specified.
        if colors is not None:
            cycle = itertools.cycle([c for c in colors])
        # Otherwise use defaults
        else:
            cycle = self.colors
            
        for linecol in linecols:
            if debug:
                print('adding line for "' + linecol + '"')

            if linecol not in data.columns:
                print('"' + linecol + '" is not in data')
                raise ValueError

            # Get default settings from stylefile
            marker = matplotlib.rcParams['lines.marker']
            markersize = matplotlib.rcParams['lines.markersize']
            markerfacecolor = matplotlib.rcParams['lines.markerfacecolor']
            markeredgecolor = matplotlib.rcParams['lines.markeredgecolor']
            markeredgewidth = matplotlib.rcParams['lines.markeredgewidth']
            _linewidth = matplotlib.rcParams['lines.linewidth']
            # If linewidth was specified as function input, use it
            if linewidth is not None:
                _linewidth = linewidth

            # If drawstyle is specified use it
            if drawstyle is not None:
                _drawstyle = drawstyle
            else:
                _drawstyle = self.drawstyle
            
            linestyle = matplotlib.rcParams['lines.linestyle']
            # Initialize color as None
            color = None
            # Add legend for this entry
            legend = True

            dashes = None
            dash_capstyle = None
            
            # Get any attributes that were assigned to this column
            if attrs is not None and linecol in attrs:
                # This should be a dict containing attributes for this column
                _attrs = attrs[linecol]
                if debug:
                    print(_attrs)

                # If any were specified, overwrite stylefile
                if 'marker' in _attrs:
                    marker = _attrs['marker']
                    
                if 'markersize' in _attrs:
                    markersize = _attrs['markersize']

                if 'markerfacecolor' in _attrs:
                    markerfacecolor = _attrs['markerfacecolor']

                if 'markeredgecolor' in _attrs:
                    markeredgecolor = _attrs['markeredgecolor']

                if 'markeredgewidth' in _attrs:
                    markeredgewidth = _attrs['markeredgewidth']
                    
                if 'linewidth' in _attrs:
                    _linewidth = _attrs['linewidth']

                if 'linestyle' in _attrs:
                    linestyle = _attrs['linestyle']
                    # If spcecial case, implement
                    if linestyle == 'imfdash':
                        linestyle = '--'
                        dashes = [10, 3]
                    elif linestyle == 'imfround':
                        linestyle = '-'
                        dashes = (0.5, 5)
                        dash_capstyle = 'round'                        

                if 'color' in _attrs:
                    color = _attrs['color']

                if 'drawstyle' in _attrs:
                    _drawstyle = _attrs['drawstyle']
                    

                if 'dashes' in _attrs:
                    dashes = _attrs['dashes']
                    
                if 'dash_capstyle' in _attrs:
                    dash_capstyle = _attrs['dash_capstyle']

                if 'legend' in _attrs:
                    legend = _attrs['legend']
            # end of linecol is in attrs

            # If color was not specified from attrs, get from color cycle
            if color is None:
                color = next(cycle)

            if marker.strip().lower() == 'none':
                marker = None

            # Make a copy of this column self.data[linecol] and decide whether to remove NA points.
            # This will prevent chart from being broken up if there are NA values.
            if linebreaks:
                # Just make a copy of the original, don't remove NA
                _df = self.data[[linecol]]
            else:
                # Remove NA points
                _df = self.data[[linecol]].dropna()
                
            if axis == 'left':
                _ax = self.ax
            elif axis == 'right':
                if self.ax_right is None:
                    self.ax_right = self.ax.twinx()
                    # Set color cycler to be common with the left y-axis.
                    # Hack from https://github.com/matplotlib/matplotlib/issues/19479
                    self.ax_right._get_lines = self.ax._get_lines
                _ax = self.ax_right
            else:
                print('axis must be left or right, given ' + str(axis))
                raise VaueError

            # Allow iteration over multiple markers into one legend
            entries = []
                
            if marker is None:
                if dashes or dash_capstyle:
                    entry = _ax.plot(_df.index, _df[linecol], label=linecol,
                                     markersize=markersize, marker=marker,
                                     markerfacecolor=markerfacecolor, markeredgecolor=markeredgecolor,
                                     markeredgewidth=markeredgewidth,
                                     drawstyle=_drawstyle,
                                     linestyle=linestyle, linewidth=_linewidth, color=color,
                                     dashes=dashes, dash_capstyle=dash_capstyle)
                else:
                    entry = _ax.plot(_df.index, _df[linecol], label=linecol,
                                     markersize=markersize, marker=marker,
                                     markerfacecolor=markerfacecolor, markeredgecolor=markeredgecolor,
                                     markeredgewidth=markeredgewidth,
                                     drawstyle=_drawstyle,
                                     linestyle=linestyle, linewidth=_linewidth, color=color)
                entries.append(entry[0])
            else:
                # Iterate over all specified markers and put in one legend entry
                for _marker in marker:
                    if dashes or dash_capstyle:
                        entry = _ax.plot(_df.index, _df[linecol], label=linecol,
                                         markersize=markersize, marker=_marker,
                                         markerfacecolor=markerfacecolor, markeredgecolor=markeredgecolor,
                                         markeredgewidth=markeredgewidth,
                                         linestyle=linestyle, linewidth=_linewidth, color=color,
                                         drawstyle=_drawstyle,
                                         dashes=dashes, dash_capstyle=dash_capstyle)
                    else:
                        entry = _ax.plot(_df.index, _df[linecol], label=linecol,
                                         markersize=markersize, marker=_marker,
                                         markerfacecolor=markerfacecolor, markeredgecolor=markeredgecolor,
                                         markeredgewidth=markeredgewidth,
                                         drawstyle=_drawstyle,
                                         linestyle=linestyle, linewidth=_linewidth, color=color)
                        
                    if legend:
                        entries.append(entry[0])
                # end of loop over marker
                
            if legend:
                self.legend_entries.append(tuple(entries))
                self.legend_labels.append(linecol)
        # end of loop over linecols

        # Re-create legend
        if self.show_legend:
            self.legend()
            if debug:
                print('called legend()')

        # Set x-axis range if specified
        if xrange is not None or margins is not None:
            if debug:
                print('before calling xrange()')
                print(xrange)
            # If margins is specified, set it for this chart and use it.
            if margins is not None:
                self.margins = margins
            self.xrange(xrange, margins=self.margins, debug=debug)
            if debug:
                print('after calling xrange():')
                print(self._xrange)

        # Set xaxis format.
        # If xformat was specified use it, otherwise use self.xformat
        if xformat is None:
            xformat = self.xformat
        self.xaxis_format(xformat=xformat, debug=debug)
        
        # Set x-axis ticks
        self.ticks(axis='x')

        # Set which x-axis is drawn on top
        self.topaxis(self._topaxis)

        # If yrange is specified use it, otherwise use self.yrange
        if axis == 'left':
            if yrange is None:
                if getattr(self, 'yrange', None) is not None:
                    yrange = self._yrange

            if yrange is not None:
                self._yrange = self._parse_yrange(yrange)
                self.yrange(self._yrange)
        if axis == 'right':
            if ryrange is None:
                if getattr(self, 'ryrange', None) is not None:
                    ryrange = self._ryrange
                    
            if ryrange is not None:
                self._ryrange = self._parse_yrange(ryrange)
                self.ryrange(self._ryrange)

    def bars(self, data=None, cols=None, indexcol=None, axis='left', colors=None,
             stack=True, total_barwidth=None, linewidth=None, edgecolor=None,
             attrs=None,
             xrange=None, margins=None, xformat=None, yrange=None, ryrange=None,
             debug=False):
        '''
        Add bar to chart
        '''

        if debug:
            print('Calling bars on "' + str(cols) + '"')

        # Set data.
        # If not specified use self.data.
        if data is None:
            data = self.data
            
        if indexcol is not None:
            if type(data) == pd.DataFrame:
                # Check if index has already been set for data
                if data.index.name == indexcol:
                    pass

                # Otherwise try to set index to indexcol
                elif indexcol not in data.columns:
                    error = 'indexcol specified as "' + str(indexcol) + '" but not found in data:\n'
                    error += str(data)
                    raise ValueError(error)
                else:
                    data.set_index(indexcol, inplace=True)
                    # If possible, convert to datetime index
                    try:
                        data.index = pd.to_datetime(data.index)
                    except Exception:
                        pass
            else:
                print('WARNING: indexcol specified but data is not DataFrame')
                
        barcols = _parse_cols(cols)
        # Don't try to plot indexcol if it was given since this is now the index
        if indexcol is not None:
            if indexcol in barcols:
                barcols.remove(indexcol)
        if debug:
            print('barcols:')
            print(barcols)

        # If no data and/or linecols are available, finish.
        if data is None:
            if barcols == []:
                print('WARNING: no data given to bars()')
            else:
                print('WARNING: no data given to bars() but barcols given as ' + str(barcols))
            print('No bars will be added')
            return

        # Set self.data to be input data
        self.data = data
        # Set self.xaxis_type from self.data
        self.xaxis_type = self._set_xaxis_type()
        if debug:
            print('self.xaxis_type = ' + str(self.xaxis_type))
            
        # Set x-axis limits if specified, or use default
        if xrange is not None:
            xrange = self._parse_xrange(xrange, debug=debug)
            self._xrange = xrange
        else:
            xrange = self._xrange
        if debug:
            print('xrange:')
            print(xrange)
            
        # Trim data as needed            
        self._trim_data(xrange, debug=debug)
        if debug:
            print(self.data)
            
        # Line width of bar borders for all bars.
        # If specified as input arg, use
        if linewidth is not None:
            _linewidth = linewidth
        # Otherwise use class settings
        else:
            _linewidth = self.linewidth
        
        # If total_barwidth is default of None,
        # set depending on self.xaxis_type
        if total_barwidth is None:
            if self.xaxis_type in ['numerical', 'categorical']:
                total_barwidth = 0.9
            elif self.xaxis_type == 'datetime':
                # Set below based on freq
                pass
            else:
                raise ValueError('total_barwdith not implemented for self.xaxis_type of ' + str(self.xaxis_type))

        # Guess freq of data and set bar width.
        # This can be overridden later for each individual bar from attrs.
        if self.xaxis_type == 'datetime':
            freq = guess_freq(self.data)
            if freq == 'D':
                barwidth = 1
            elif freq == 'W':
                barwidth = 3
            elif freq == 'M':
                barwidth = 20
            elif freq == 'Q':
                barwidth = 70
            elif freq == 'A':
                barwidth =300
            elif freq == '?':
                barwidth = 1
            else:
                raise ValueError('Unknown frequency ' + str(freq))

            # If stack=False, need to divide each barwidth by number of bars.
            if not stack:
                if total_barwidth is None:
                    # Set to total barwdith
                    total_barwidth = barwidth
                    
                # Set each bar to be 1 / len(barcols)
                barwidth /= len(barcols)
                
        elif self.xaxis_type in ['categorical', 'numerical']:
            # Split barwidth among barcols.
            # This can be overridden later for each individual bar from attrs.
            barwidth = total_barwidth / len(barcols)
        else:
            raise ValueError('self.xaxis_type of ' + str(self.xaxis_type) + ' not implemented')

        # Iterate over barcols and plot each bar.
        # Create cycle if specified.
        if colors is not None:
            cycle = itertools.cycle([c for c in colors])
        # Otherwise use defaults
        else:
            cycle = self.colors
        if debug:
            print('cycle:')
            print(cycle)
            
        # Need to keep track of positive and negative offsets if stacked.
        pos_offset = [0] * len(self.data)
        neg_offset = [0] * len(self.data)

        # For stack=False, need offset for x-axis.
        # Default is to start from half the total barwidth and half a bar width in the negative direction,
        # this centers the bars.
        if not stack:
            total_offset = - total_barwidth / 2. - total_barwidth / len(barcols) / 2.

        # Keep track of all patches from previous barcols so that
        # if individual bar colors are specified for a barcol,
        # we do not color other bars.
        previous_patches = []
            
        for ibarcol, barcol in enumerate(barcols):
            if debug:
                print('-' * 40)
                print('barcol = "' + barcol + '"')
                
            if barcol not in data.columns:
                print('"' + barcol + '" is not in data')
                raise ValueError

            # Get default settings from stylefile
            # The patch.edgecolor is used for both the hatch line color
            # and the bar edge color.
            # For IMF charts use this as the hatch line color
            # and draw bar again with edgecolor set to black.
            hatchcolor = matplotlib.rcParams['patch.edgecolor']
            hatch = None
            hatchwidth = matplotlib.rcParams['hatch.linewidth']
            # Local copy for this column, can be overwritten with attrs
            # If edgecolor has not been set, use default
            if edgecolor is None:
                _edgecolor = self.baredgecolor
            else:
                _edgecolor = edgecolor

            # Individual bar colors
            barcolors = None

            # Add legend for this entry
            legend = True

            # Flag for whether offset was specified for this column.
            # If not specified, default to fixed offset for each col in barcols.
            offset_specified = False

            # Get any attributes that were assigned to this column
            if attrs is not None and barcol in attrs:
                # This should be a dict containing attributes for this column
                _attrs = attrs[barcol]
                if debug:
                    print(_attrs)

                # If any were specified, overwrite stylefile

                # Hatch colors
                if 'hatchcolor' in _attrs:
                    hatchcolor = _attrs['hatchcolor']

                # Hatch pattern in bars
                if 'hatch' in _attrs:
                    hatch = _attrs['hatch']

                # Line width of hatches
                if 'hatchwidth' in _attrs:
                    hatchwidth = _attrs['hatchwidth']

                # Color of bar edges
                if 'edgecolor' in _attrs:
                    _edgecolor = _attrs['edgecolor']
                    
                # If offset is specified for when stack=False, use it
                if 'offset' in _attrs:
                    offset = _attrs['offset']
                    offset_specified = True

                # If barwidth is specified, use it
                if 'width' in _attrs:
                    width = _attrs['width']
                    
                # Add legend entry
                if 'legend' in _attrs:
                    legend = _attrs['legend']

                # Get individual bar colors
                if 'barcolors' in _attrs:
                    barcolors = _attrs['barcolors']

            # end of attrs existing for this barcol

            # Set offset for this bar.
            # Whether the offset is specified or not,
            # the barwidth is added into the offset.
            # Since all offses are cumulative, adding an offset
            # to the first bar will just shift all bars.
            # Adding additional offsets to other bars will
            # add spacing in between bars.
            if not offset_specified:
                offset = barwidth
            else:
                offset += barwidth
                
            # For color, if it is specified use it,
            # otherwise get the next color from the color cycle.
            if attrs is not None and barcol in attrs and 'color' in attrs[barcol]:
                color = attrs[barcol]['color']
                if debug:
                    print('1. setting color for ' + barcol + ' to ' + color)
            else:
                color = next(cycle)

            if debug:
                if attrs is not None and barcol in attrs:
                    print('attrs:')
                    print(attrs[barcol])
                print('hatch = ' + str(hatch))
                print('linewidth = ' + str(_linewidth))
                print('_edgecolor = ' + str(_edgecolor))
                print('hatchcolor = ' + str(hatchcolor))
                print('color = ' + str(color))

            # Make copy of positive and negative parts.
            # Set mask to NA so that they don't show up with lines in the chart,
            # but when saving to pos_offset and neg_offset, set values to 0.
            _df_pos = self.data[[barcol]].copy()
            mask = _df_pos[barcol] < 0
            _df_pos.loc[mask, barcol] = np.nan
                
            _df_neg = self.data[[barcol]].copy()
            mask = _df_neg[barcol] > 0
            _df_neg.loc[mask, barcol] = np.nan

            # Set offset when stack=False
            if not stack:
                total_offset += offset
                # If stack=False, need to set _x to be coordinates
                # of where bars are.
                if self.xaxis_type == 'categorical':
                    _x = np.arange(len(self.data)) + total_offset
                else:
                    _x = self.data.index + pd.Timedelta(days=total_offset)

                if debug:
                    print('-' * 40)
                    print(barcol)
                    print('barwidth = ' + str(barwidth) + ' offset = ' + str(offset))
                    print('_x:')
                    print(_x)
            
            if axis == 'left':
                _ax = self.ax
            elif axis == 'right':
                if self.ax_right is None:
                    self.ax_right = self.ax.twinx()
                    # Set color cycler to be common with the left y-axis.
                    # Hack from https://github.com/matplotlib/matplotlib/issues/19479
                    self.ax_right._get_lines = self.ax._get_lines
                _ax = self.ax_right
            else:
                print('axis must be left or right, given ' + str(axis))
                raise VaueError
                
            if stack:
                # No direct way to set hatch line widths in ax.bar,
                # need to use plt.rc_context()
                with plt.rc_context({"hatch.linewidth": hatchwidth}):
                    if ibarcol == 0:
                        # Draw negative first
                        entry = _ax.bar(self.data.index, _df_neg[barcol],
                                        width=barwidth, color=color, edgecolor=hatchcolor, hatch=hatch, linewidth=0,
                                        zorder=1, label=barcol)
                        # Draw positive
                        entry = _ax.bar(self.data.index, _df_pos[barcol],
                                        width=barwidth, color=color, edgecolor=hatchcolor, hatch=hatch, linewidth=0,
                                        zorder=1, label=barcol)
                    else:
                        # Draw negative first
                        entry = _ax.bar(self.data.index, _df_neg[barcol],
                                        bottom=neg_offset,
                                        width=barwidth, color=color, edgecolor=hatchcolor, hatch=hatch, linewidth=0,
                                        zorder=1, label=barcol)
                        # Draw positive
                        entry = _ax.bar(self.data.index, _df_pos[barcol],
                                        bottom=pos_offset,
                                        width=barwidth, color=color, edgecolor=hatchcolor, hatch=hatch, linewidth=0,
                                        zorder=1, label=barcol)
                            
                    # Draw again with _edgecolor
                    entry2 = _ax.bar(self.data.index, _df_pos[barcol],
                                     width=barwidth, color='none', edgecolor=_edgecolor, linewidth=_linewidth,
                                     bottom=pos_offset,
                                     zorder=2)
                    _ = _ax.bar(self.data.index, _df_neg[barcol],
                                bottom=neg_offset,
                                width=barwidth, color='none', edgecolor=_edgecolor, linewidth=_linewidth,
                                zorder=2)
                    
                    # If barcolors was specified, get all patches for this barcol,
                    # and if the index matches, set color for the corresponding bar.
                    patches = [p for p in _ax.patches if p not in previous_patches]
                    if debug:
                        print('len(patches) = ' + str(len(patches)))

                    # Add individually specified colors
                    if barcolors is not None:
                        color_bars(barcolors, patches, self.data.index, stack=True, debug=debug)
                        
                    # Add current patches to previous_patches
                    previous_patches += patches
                # end of plt.rc_context()
            # end of stack
            else:
                entry = _ax.bar(_x, self.data[barcol],
                                width=barwidth, color=color, edgecolor=hatchcolor, hatch=hatch, linewidth=0,
                                zorder=1, label=barcol)
                
                # Draw again with _edgecolor
                entry2 = _ax.bar(_x, self.data[barcol],
                                 width=barwidth, color='none', edgecolor=_edgecolor, linewidth=_linewidth,
                                 zorder=2)
                    
                # If barcolors was specified, get all patches for this barcol,
                # and if the index matches, set color for the corresponding bar.
                patches = [p for p in _ax.patches if p not in previous_patches]
                if debug:
                    print('len(patches) = ' + str(len(patches)))

                # Add individually specified colors
                if barcolors is not None:
                    color_bars(barcolors, patches, self.data.index, stack=False, debug=debug)
                        
                # Add current patches to previous_patches
                previous_patches += patches
                    
            # end of stack=False
            
            # Adjust offsets
            neg_offset += _df_neg[barcol].replace(np.nan, 0).values
            pos_offset += _df_pos[barcol].replace(np.nan, 0).values

            if legend:
                # Combine entries for entry without and with border line
                self.legend_entries.append(tuple([entry[0], entry2[0]]))
                self.legend_labels.append(barcol)
        # end of loop over barcols

        # Re-create legend
        if self.show_legend:
            self.legend()
        
        # Set x-axis range if specified
        if xrange is not None or margins is not None:
            if debug:
                print('before calling xrange()')
                print(xrange)
            # If margins is specified, set it for this chart and use it.
            if margins is not None:
                self.margins = margins
            self.xrange(xrange, margins=self.margins, debug=debug)
            if debug:
                print('after calling xrange():')
                print(self._xrange)

        # Set xaxis format.
        # If xformat was specified use it, otherwise use self.xformat
        if xformat is None:
            xformat = self.xformat
        self.xaxis_format(xformat=xformat, debug=debug)
        
        # Set x-axis ticks
        self.ticks(axis='x')

        # Set which x-axis is drawn on top
        self.topaxis(self._topaxis)

        # If yrange is specified use it, otherwise use self.yrange
        if axis == 'left':
            if yrange is None:
                if getattr(self, '_yrange', None) is not None:
                    yrange = self._yrange

            if yrange is not None:
                self._yrange = self._parse_yrange(yrange)
                self.yrange(self._yrange)
        if axis == 'right':
            if ryrange is None:
                if getattr(self, '_ryrange', None) is not None:
                    ryrange = self._ryrange
                    
            if ryrange is not None:
                self._ryrange = self._parse_yrange(ryrange)
                self.ryrange(self._ryrange)

    def area(self, data=None, cols=None, indexcol=None, axis='left', colors=None, alpha=1,
             stack=True, linewidth=None, edgecolor=None,
             attrs=None,
             xrange=None, margins=0, xformat=None, yrange=None, ryrange=None,
             debug=False):
        '''
        Add area to chart
        '''

        if debug:
            print('Calling area on "' + str(cols) + '"')
            
        # Set data.
        # If not specified use self.data.
        if data is None:
            data = self.data
            
        if indexcol is not None:
            if type(data) == pd.DataFrame:
                # Check if index has already been set for data
                if data.index.name == indexcol:
                    pass

                # Otherwise try to set index to indexcol
                elif indexcol not in data.columns:
                    error = 'indexcol specified as "' + str(indexcol) + '" but not found in data:'
                    error += str(data)
                    raise ValueError(error)
                else:
                    data.set_index(indexcol, inplace=True)
                    # If possible, convert to datetime index
                    try:
                        data.index = pd.to_datetime(data.index)
                    except Exception:
                        pass
            else:
                print('WARNING: indexcol specified but data is not DataFrame')
                
        areacols = _parse_cols(cols)
        # Don't try to plot indexcol if it was given since this is now the index
        if indexcol is not None:
            if indexcol in areacols:
                areacols.remove(indexcol)
        if debug:
            print('areacols:')
            print(areacols)

        # If no data and/or areacols are available, finish.
        if data is None:
            if areacols == []:
                print('WARNING: no data given to areas()')
            else:
                print('WARNING: no data given to areas() but areacols given as ' + str(areacols))
            print('No areas will be added')
            return

        # Line width of area borders for all areas.
        # If specified as input arg, use
        if linewidth is not None:
            _linewidth = linewidth
        # Otherwise use class settings
        else:
            _linewidth = self.arealinewidth
            
        # Set self.data to be input data
        self.data = data
        # Set self.xaxis_type from self.data
        self.xaxis_type = self._set_xaxis_type()
        if debug:
            print('self.xaxis_type = ' + str(self.xaxis_type))
            
        # Set x-axis limits if specified, or use default
        if xrange is not None or margins is not None:
            xrange = self._parse_xrange(xrange, debug=debug)
            self._xrange = xrange
        else:
            xrange = self._xrange
        if debug:
            print('xrange:')
            print(xrange)
            
        # Trim data as needed            
        self._trim_data(xrange, debug=debug)
        if debug:
            print(self.data)

        # Iterate over areacols and plot each area.
        # Create cycle if specified.
        if colors is not None:
            cycle = itertools.cycle([c for c in colors])
        # Otherwise use defaults
        else:
            cycle = self.colors
        if debug:
            print('cycle:')
            print(cycle)
            
        # Need to keep track of positive and negative offsets if stacked.
        pos_offset = [0] * len(self.data)
        neg_offset = [0] * len(self.data)

        # Keep track of whether any negative area exists.
        # If no negative area exists, then set min yrange to 0 at end.
        has_negative = False

        for iareacol, areacol in enumerate(areacols):
            if debug:
                print('-' * 40)
                print('areacol = "' + areacol + '"')
                
            if areacol not in data.columns:
                print('"' + areacol + '" is not in data')
                raise ValueError

            # Get default settings from stylefile
            # The patch.edgecolor is used for both the hatch line color
            # and the area edge color.
            # For IMF charts use this as the hatch line color
            # and draw area again with edgecolor set to black.
            hatchcolor = matplotlib.rcParams['patch.edgecolor']
            hatch = None
            hatchwidth = matplotlib.rcParams['hatch.linewidth']
            # Local copy for this column, can be overwritten with attrs
            # If areaedgecolor has not been set, use default
            if edgecolor is None:
                _edgecolor = self.areaedgecolor
            else:
                _edgecolor = edgecolor

            # Add legend for this entry
            legend = True

            # alpha
            if alpha is None:
                _alpha = self.alpha
            else:
                _alpha = alpha

            # Get any attributes that were assigned to this column
            if attrs is not None and areacol in attrs:
                # This should be a dict containing attributes for this column
                _attrs = attrs[areacol]
                if debug:
                    print(_attrs)

                # If any were specified, overwrite stylefile

                # Hatch colors
                if 'hatchcolor' in _attrs:
                    hatchcolor = _attrs['hatchcolor']

                # Hatch pattern in area
                if 'hatch' in _attrs:
                    hatch = _attrs['hatch']

                # Line width of hatches
                if 'hatchwidth' in _attrs:
                    hatchwidth = _attrs['hatchwidth']

                # Color of area edges
                if 'edgecolor' in _attrs:
                    _edgecolor = _attrs['edgecolor']

                # alpha
                if 'alpha' in _attrs:
                    _alpha = _attrs['alpha']
                    
                # Add legend entry
                if 'legend' in _attrs:
                    legend = _attrs['legend']

            # end of attrs existing for this areacol

            # For color, if it is specified use it,
            # otherwise get the next color from the color cycle.
            if attrs is not None and areacol in attrs and 'color' in attrs[areacol]:
                color = attrs[areacol]['color']
                if debug:
                    print('1. setting color for ' + areacol + ' to ' + color)
            else:
                color = next(cycle)

            if debug:
                if attrs is not None and areacol in attrs:
                    print('attrs:')
                    print(attrs[areacol])
                print('hatch = ' + str(hatch))
                print('_edgecolor = ' + str(_edgecolor))
                print('hatchcolor = ' + str(hatchcolor))
                print('color = ' + str(color))

            # Make copy of positive and negative parts
            # (for area chart we should not have negative parts).
            # Set mask to NA so that they don't show up with lines in the chart,
            # but when saving to pos_offset and neg_offset, set values to 0.
            _df_pos = self.data[[areacol]].copy()
            mask = _df_pos[areacol] < 0
            _df_pos.loc[mask, areacol] = np.nan
                
            _df_neg = self.data[[areacol]].copy()
            mask = _df_neg[areacol] > 0
            _df_neg.loc[mask, areacol] = np.nan

            if len(_df_neg[areacol].dropna()) > 0:
                has_negative = True

            if axis == 'left':
                _ax = self.ax
            elif axis == 'right':
                if self.ax_right is None:
                    self.ax_right = self.ax.twinx()
                    # Set color cycler to be common with the left y-axis.
                    # Hack from https://github.com/matplotlib/matplotlib/issues/19479
                    self.ax_right._get_lines = self.ax._get_lines
                _ax = self.ax_right
            else:
                print('axis must be left or right, given ' + str(axis))
                raise VaueError
                    
            if stack:
                # No direct way to set hatch line widths in ax.area,
                # need to use plt.rc_context()
                with plt.rc_context({"hatch.linewidth": hatchwidth}):
                    # Draw without line edges
                    # Draw negative first
                    entry = _ax.fill_between(self.data.index, y1=_df_neg[areacol] + neg_offset, y2=neg_offset,
                                             color=color, edgecolor=hatchcolor, hatch=hatch, linewidth=0, alpha=_alpha,
                                             zorder=1, label=areacol)
                    # Draw positive
                    entry = _ax.fill_between(self.data.index, y1=_df_pos[areacol] + pos_offset, y2=pos_offset,
                                             color=color, edgecolor=hatchcolor, hatch=hatch, linewidth=0, alpha=_alpha,
                                             zorder=1, label=areacol)
                        
                    # Draw again with _edgecolor
                    # positive
                    entry2 = _ax.fill_between(self.data.index, y1=_df_pos[areacol] + pos_offset, y2=pos_offset,
                                              color='none', edgecolor=_edgecolor, linewidth=_linewidth, alpha=_alpha,
                                              zorder=2)
                    # negative
                    _ = _ax.fill_between(self.data.index, y1=_df_neg[areacol] + neg_offset, y2=neg_offset,
                                         color='none', edgecolor=_edgecolor, linewidth=_linewidth, alpha=_alpha,
                                         zorder=2)

                # end of plt.rc_context()
            # end of stack
            else:
                # Draw once without edgecolor
                entry = _ax.fill_between(self.data.index, y1=self.data[areacol], y2=0,
                                         color=color, edgecolor=hatchcolor, hatch=hatch, linewidth=0, alpha=_alpha,
                                         zorder=1, label=areacol)
                
                # Draw again with _edgecolor
                entry2 = _ax.fill_between(self.data.index, y1=self.data[areacol], y2=0,
                                          color='none', edgecolor=_edgecolor, linewidth=_linewidth, alpha=_alpha,
                                          zorder=2)
            # end of stack=False
            
            # Adjust offsets
            neg_offset += _df_neg[areacol].replace(np.nan, 0).values
            pos_offset += _df_pos[areacol].replace(np.nan, 0).values

            if debug:
                print('pos_offset:')
                print(pos_offset)
                print('neg_offset:')
                print(neg_offset)

            # If adding to legend
            if legend:
                # Create a Patch that has the same characteristics
                # that can be added to the legend entries.
                entry = Patch(
                    facecolor=color,
                    edgecolor=_edgecolor,
                    hatch=hatch,
                    alpha=_alpha,
                    label=areacols)
                self.legend_entries.append(entry)
                self.legend_labels.append(areacol)
        # end of loop over areacols

        # Re-create legend
        if self.show_legend:
            self.legend()
        
        # Set x-axis range if specified.
        # For area, default is to set margins=0 so that the are chart
        # does not look like it was chopped off at the edges.
        # To have same behavior as lines etc., use margins='auto'.
        if xrange is not None or margins is not None:
            if debug:
                print('before calling xrange()')
                print(xrange)
            # If margins is specified, set it for this chart and use it.
            if margins is not None:
                self.margins = margins
            self.xrange(xrange, margins=self.margins, debug=debug)
            if debug:
                print('after calling xrange():')
                print(self._xrange)

        # Set xaxis format.
        # If xformat was specified use it, otherwise use self.xformat
        if xformat is None:
            xformat = self.xformat
        self.xaxis_format(xformat=xformat, debug=debug)
        
        # Set x-axis ticks
        self.ticks(axis='x')

        # Set which x-axis is drawn on top
        self.topaxis(self._topaxis)

        # If yrange is specified use it, otherwise use self.yrange
        if axis == 'left':
            if yrange is None:
                if getattr(self, '_yrange', None) is not None:
                    yrange = self._yrange

            if yrange is not None:
                self._yrange = self._parse_yrange(yrange)
                self.yrange(self._yrange)

            # If no yrange is given and data is all above 0,
            # set y-axis min to 0
            if yrange is None and not has_negative:
                self.ax.set_ylim(bottom=0)
                
        if axis == 'right':
            if ryrange is None:
                if getattr(self, '_ryrange', None) is not None:
                    ryrange = self._ryrange
                    
            if ryrange is not None:
                self._ryrange = self._parse_yrange(ryrange)
                self.ryrange(self._ryrange)

            # If no yrange is given and data is all above 0,
            # set y-axis min to 0
            if ryrange is None and not has_negative:
                self.ax_right.set_ylim(bottom=0)
                
    def scatter(self, data, cols, indexcol=None, attrs=None, debug=False):
        '''
        Add scatter to chart
        '''

        if debug:
            print('Calling scatter on "' + str(cols) + '"')

        if indexcol is not None:
            if type(data) == pd.DataFrame:
                # Check if index has already been set for data
                if data.index.name == indexcol:
                    pass

                # Otherwise try to set index to indexcol
                elif indexcol not in data.columns:
                    error = 'indexcol specified as "' + str(indexcol) + '" but not found in data:\n'
                    error += str(data)
                    raise ValueError(error)
                else:
                    data.set_index(indexcol, inplace=True)
            else:
                print('WARNING: indexcol specified but data is not DataFrame')

    def topaxis(self, topaxis='left'):
        '''
        Set which x-axis is drawn on top.
        '''

        if topaxis == 'left':
            if self.ax_right:
                self.ax.set_zorder(self.ax_right.get_zorder() + 1)
                self.ax.patch.set_visible(False)
        elif topaxis == 'right':
            if self.ax_right:
                self.ax_right.set_zorder(self.ax.get_zorder() + 1)
                self.ax_right.patch.set_visible(False)
        else:
            print('WARNING:')
            print('Chart.topaxis must be "left" or "right", given ' + str(topaxis))

    def hline(self, y, xrange=None, coordinates='data', color='red', linewidth=1, linestyle='-', alpha=1, dashes=None, dash_capstyle=None,
              label='', legend=False, zorder=3,
              debug=False, **kwarg):
        '''
        Add horizontal line across figure.

        If xrange is None, a horizontal line will be drawn across the entire chart.
        Otherwise range should be a list of two values or a str that has form "start:end".

        `coordinates` is either "data" or "axis", for "data" xrange is converted to data coordinates.
        For "axis" the fraction of the self.ax is used.
        '''

        xrange = self._parse_xrange(xrange)
        if debug:
            print('xrange:')
            print(xrange)
            
        # No xrange
        if xrange == [None, None]:
            if debug:
                print('adding hline with no xrange')
            # dashes or dash_capstyle specified
            if dashes or dash_capstyle:
                entry = self.ax.axhline(y=y, color=color, linewidth=linewidth, linestyle=linestyle, alpha=alpha,
                                        dashes=dashes, dash_capstyle=dash_capstyle, zorder=zorder)
            # no dashes or dash_capstyle specified
            else:
                entry = self.ax.axhline(y=y, color=color, linewidth=linewidth, linestyle=linestyle, alpha=alpha, zorder=zorder)

        # xrange given
        else:
            if debug:
                print('adding hline with xrange')

            # If xrange is given as Timestamps, need to calculate fraction of x-axis range.
            if coordinates == 'data' and type(xrange[0]) == pd.Timestamp and type(xrange[1]) == pd.Timestamp:
                # Get axis range, this will be in ordinals
                xmin, xmax = self.ax.get_xlim()
                # Get fraction of xrange[0], xrange[1]
                xrange[0] = (xrange[0] - pd.Timestamp('1970-01-01')).days
                xrange[1] = (xrange[1] - pd.Timestamp('1970-01-01')).days
                             
                xrange[0] = (xrange[0] - xmin) / (xmax - xmin)
                xrange[1] = (xrange[1] - xmin) / (xmax - xmin)
                if debug:
                    print('xrange:')
                    print(xrange)
                
            # dashes or dash_capstyle specified
            if dashes or dash_capstyle:
                entry = self.ax.axhline(y=y, xmin=xrange[0], xmax=xrange[1],
                                        color=color, linewidth=linewidth, linestyle=linestyle, alpha=alpha,
                                        dashes=dashes, dash_capstyle=dash_capstyle, zorder=zorder)
            # no dashes or dash_capstyle specified
            else:
                entry = self.ax.axhline(y=y, xmin=xrange[0], xmax=xrange[1],
                                        color=color, linewidth=linewidth, linestyle=linestyle, alpha=alpha, zorder=zorder)

        # If adding to legend
        if legend:
            self.legend_entries.append(entry)
            self.legend_labels.append(label)

        # Re-create legend
        if self.show_legend:
            self.legend()
            if debug:
                print('called legend()')
            
    def vline(self, x, yrange=None, coordinates='data', width=1, color='red', linewidth=1, linestyle='-', alpha=1, dashes=None, dash_capstyle=None,
              label='', legend=False, zorder=3,
              debug=False, **kwarg):
        '''
        Add vertical line across figure.
        If yrange is None, a vertical line will be drawn across the entire chart.
        Otherwise range should be a list of two values.
        
        `coordinates` is either "data" or "axis", for "data" xrange is converted to data coordinates.
        For "axis" the fraction of the self.ax is used.
        '''

        yrange = self._parse_yrange(yrange)

        # If x-axis is datetime, make sure to convert x to datetime
        if self.xaxis_type == 'datetime':
            try:
                x = pd.Timestamp(x)
            except ValueError:
                print('Could not convert ' + str(x) + ' to pd.Timestamp')
                raise ValueError

        # Convert yrange to fraction of self.ax range as needed
        if coordinates == 'data' and yrange != [None, None]:
            # Get axis range, this will be in ordinals
            ymin, ymax = self.ax.get_ylim()

            # Get fraction of xrange[0], xrange[1]
            yrange[0] = (yrange[0] - ymin) / (ymax - ymin)
            yrange[1] = (yrange[1] - ymin) / (ymax - ymin)

        # No yrange
        if yrange == [None, None]:
            # dashes or dash_capstyle specified
            if dashes or dash_capstyle:
                entry = self.ax.axvline(x=x, color=color, linewidth=linewidth, linestyle=linestyle, dashes=dashes, alpha=alpha,
                                        dash_capstyle=dash_capstyle, zorder=zorder)
                # no dashes or dash_capstyle specified
            else:
                entry = self.ax.axvline(x=x, color=color, linewidth=linewidth, linestyle=linestyle, alpha=alpha, zorder=zorder)

        # yrange given
        else:
            # dashes or dash_capstyle specified
            if dashes or dash_capstyle:
                entry = self.ax.axvline(x=x, ymin=yrange[0], ymax=yrange[1],
                                        color=color, linewidth=linewidth, linestyle=linestyle, alpha=alpha,
                                        dashes=dashes, dash_capstyle=dash_capstyle, zorder=zorder)
                # no dashes or dash_capstyle specified
            else:
                entry = self.ax.axvline(x=x, ymin=yrange[0], ymax=yrange[1],
                                        color=color, linewidth=linewidth, linestyle=linestyle, alpha=alpha, zorder=zorder)


        # If adding to legend
        if legend:
            self.legend_entries.append(entry)
            self.legend_labels.append(label)

        # Re-create legend
        if self.show_legend:
            self.legend()
            if debug:
                print('called legend()')
                
    def hrect(self, ymin=0, ymax=0, xrange=None, coordinates='data', color='red', linecolor='none', linewidth=0, linestyle='-', alpha=0.3,
              dash_capstyle=None,
              hatch=None, # hatchlinewidth=None,
              label='', legend=False,
              zorder=0,
              debug=False, **kwarg):
        '''
        Add horizontal rectangle across figure.

        If xrange is None, a horizontal rectangle will be drawn across the entire chart.
        Otherwise range should be a list of two values or a str that has form "start:end".

        `coordinates` is either "data" or "axis", for "data" xrange is converted to data coordinates.
        For "axis" the fraction of the self.ax is used.
        '''

        xrange = self._parse_xrange(xrange)
        if debug:
            print('xrange:')
            print(xrange)
            
        # No xrange
        if xrange == [None, None]:
            if debug:
                print('adding rect with no xrange')
            xrange = [0, 1]
        # xrange given
        else:
            if debug:
                print('adding hline with xrange')

            # If xrange is given as Timestamps, need to calculate fraction of x-axis range.
            if coordinates == 'data' and type(xrange[0]) == pd.Timestamp and type(xrange[1]) == pd.Timestamp:
                # Get axis range, this will be in ordinals
                xmin, xmax = self.ax.get_xlim()
                # Get fraction of xrange[0], xrange[1]
                xrange[0] = (xrange[0] - pd.Timestamp('1970-01-01')).days
                xrange[1] = (xrange[1] - pd.Timestamp('1970-01-01')).days
                             
                xrange[0] = (xrange[0] - xmin) / (xmax - xmin)
                xrange[1] = (xrange[1] - xmin) / (xmax - xmin)
                if debug:
                    print('xrange:')
                    print(xrange)

        # dash_capstyle specified
        if dash_capstyle:
            entry = self.ax.axhspan(ymin, ymax, xmin=xrange[0], xmax=xrange[1],
                                    facecolor=color, linewidth=linewidth, edgecolor=linecolor, linestyle=linestyle, alpha=alpha,
                                    hatch=hatch,
                                    zorder=zorder,
                                    dash_capstyle=dash_capstyle)
        # no dash_capstyle specified
        else:
            entry = self.ax.axhspan(ymin, ymax, xmin=xrange[0], xmax=xrange[1],
                                    facecolor=color, linewidth=linewidth, edgecolor=linecolor, linestyle=linestyle, alpha=alpha,
                                    hatch=hatch,
                                    zorder=zorder)

        # If adding to legend
        if legend:
            # Create a Patch that has the same characteristics
            # that can be added to the legend entries.
            entry = Patch(
                facecolor=color,
                edgecolor=linecolor,
                hatch=hatch,
                alpha=alpha,
                label=label)
            self.legend_entries.append(entry)
            self.legend_labels.append(label)

        # Re-create legend
        if self.show_legend:
            self.legend()
            if debug:
                print('called legend()')
            
#        # As of Matplotlib 3.8.0, changing the hatch linewidth does not work.
#        # It is possible to set the rcParams temporarily,
#        # but once it is restored to the original value that value takes over.
#        # Below is code in case this is fixed in the future.
#
#        # If hatch_linewdith is set, use it.
#        # Otherwise default to rcParams
#        hatchlinewidth_rc = matplotlib.rcParams["hatch.linewidth"]
#        if hatchlinewidth is None:
#            hatchlinewidth = hatchlinewidth_rc
#
#        # The following try-finally is so that matplotlib.rcParams["hatch.linewidth"]
#        # is always resstored to original value.
#        try:
#            # Set value to hatchlinewidth
#            matplotlib.rcParams["hatch.linewidth"] = hatchlinewidth
#
#            # Add ax.axhspan here
#        finally:
#            # Set mpl.rcParams back to original value.
#            # For Matplotlib 3.8.0, this negates the effect of setting hatch.linewidth
#            # and the resulting figure will always have the hatch linewidth of rcParams.
#            matplotlib.rcParams["hatch.linewidth"] = hatchlinewidth_rc

    def vrect(self, xmin=0, xmax=0, yrange=None, coordinates='data', color='red', linecolor='none', linewidth=0, linestyle='-', alpha=0.3,
              dash_capstyle=None,
              hatch=None, # hatchlinewidth=None,
              label='', legend=False,
              zorder=0,
              debug=False, **kwarg):
        '''
        Add vertical rectangle across figure.

        If yrange is None, a vertical rectangle will be drawn across the entire chart.
        Otherwise range should be a list of two values or a str that has form "start:end".

        `coordinates` is either "data" or "axis", for "data" xrange is converted to data coordinates.
        For "axis" the fraction of the self.ax is used.
        '''

        yrange = self._parse_yrange(yrange)
        if debug:
            print('yrange:')
            print(yrange)
            
        # No yrange
        if yrange == [None, None]:
            if debug:
                print('adding rect with no yrange')
            yrange = [0, 1]
        # yrange given
        else:
            if debug:
                print('adding hline with yrange')

            # If yrange is given, need to calculate fraction of x-axis range.
            if coordinates == 'data':
                # Get axis range
                _ymin, _ymax = self.ax.get_ylim()
                # Get fraction of yrange[0], yrange[1]
                yrange[0] = (yrange[0] - _ymin) / (_ymax - _ymin)
                yrange[1] = (yrange[1] - _ymin) / (_ymax - _ymin)
                if debug:
                    print('yrange:')
                    print(yrange)

        # Parse x-axis range
        xmin, xmax = self._parse_xrange([xmin, xmax])

        # dash_capstyle specified
        if dash_capstyle:
            entry = self.ax.axvspan(xmin, xmax, ymin=yrange[0], ymax=yrange[1],
                                    facecolor=color, linewidth=linewidth, edgecolor=linecolor, linestyle=linestyle, alpha=alpha,
                                    hatch=hatch,
                                    zorder=zorder,
                                    dash_capstyle=dash_capstyle)
        # no dash_capstyle specified
        else:
            entry = self.ax.axvspan(xmin, xmax, ymin=yrange[0], ymax=yrange[1],
                                    facecolor=color, linewidth=linewidth, edgecolor=linecolor, linestyle=linestyle, alpha=alpha,
                                    hatch=hatch,
                                    zorder=zorder)
            
        # If adding to legend
        if legend:
            # Create a Patch that has the same characteristics
            # that can be added to the legend entries.
            entry = Patch(
                facecolor=color,
                edgecolor=linecolor,
                hatch=hatch,
                alpha=alpha,
                label=label)
            self.legend_entries.append(entry)
            self.legend_labels.append(label)

        # Re-create legend
        if self.show_legend:
            self.legend()
            if debug:
                print('called legend()')
            
#        # As of Matplotlib 3.8.0, changing the hatch linewidth does not work.
#        # It is possible to set the rcParams temporarily,
#        # but once it is restored to the original value that value takes over.
#        # Below is code in case this is fixed in the future.
#
#        # If hatch_linewdith is set, use it.
#        # Otherwise default to rcParams
#        hatchlinewidth_rc = matplotlib.rcParams["hatch.linewidth"]
#        if hatchlinewidth is None:
#            hatchlinewidth = hatchlinewidth_rc
#
#        # The following try-finally is so that matplotlib.rcParams["hatch.linewidth"]
#        # is always resstored to original value.
#        try:
#            # Set value to hatchlinewidth
#            matplotlib.rcParams["hatch.linewidth"] = hatchlinewidth
#
#            # Add ax.axhspan here
#        finally:
#            # Set mpl.rcParams back to original value.
#            # For Matplotlib 3.8.0, this negates the effect of setting hatch.linewidth
#            # and the resulting figure will always have the hatch linewidth of rcParams.
#            matplotlib.rcParams["hatch.linewidth"] = hatchlinewidth_rc

    def text(self, x, y, text='', color='black',
             fontsize=14, fontfamily='Segoe UI', fontweight='normal',
             va='top', ha='left', zorder=1, **kwargs):
        '''
        Add text.

        va='top' will align the top of the text to the specified value.
        ha='left' will align the left of the text to the specified value.
        '''

        if self.xaxis_type == 'datetime':
            try:
                x = pd.Timestamp(x)
            except Exception:
                print('WARNING: Could not convert ' + str(x) + ' to pd.Timestamp')

        if self.fontname is not None:
            handle = self.ax.text(x, y, text,
                                  color=color,
                                  fontsize=fontsize, fontfamily=fontfamily, fontweight=fontweight,
                                  va=va, ha=ha, zorder=zorder,
                                  fontname=self.fontname)
        else:
            handle = self.ax.text(x, y, text,
                                  color=color,
                                  fontsize=fontsize, fontfamily=fontfamily, fontweight=fontweight,
                                  va=va, ha=ha, zorder=zorder)

        # Save the handle so it can be accessed later
        self.texts.append(handle)
    
    def arrow(self, head=(0, 0), tail=(1, 1), coords='data',
              color='black', edgecolor=None, edgewidth=0,
              width=4, headwidth=15, headlength=15, shrink=0.05,
              zorder=1, **kwargs):
        '''
        Add arrow and text.
        '''

        if self.xaxis_type == 'datetime' and coords=='data':
            try:
                head[0] = pd.Timestamp(head[0])
            except Exception:
                print('WARNING: Could not convert ' + str(head[0]) + ' to pd.Timestamp')

            try:
                tail[0] = pd.Timestamp(tail[0])
            except Exception:
                print('WARNING: Could not convert ' + str(tail) + ' to pd.Timestamp')
                
        self.ax.annotate(xy=head, xytext=tail,
                         arrowprops=dict(
                             facecolor=color,      # Color of the arrow
                             edgecolor=edgecolor,  # Color of arrow outline
                             linewidth=edgewidth,  # Width of arrow edges
                             shrink=shrink,        # Distance from point and text
                             width=width,          # Width of the arrow tail in points
                             headwidth=headwidth,  # Width of the arrow head base in points
                             headlength=headlength # Length of the arrow head in points
                         ),
                         text='', xycoords=coords, textcoords=coords, zorder=zorder)

    def fill(self, lo, hi, data=None, indexcol=None, axis='left',
             color='red', linecolor='none', linewidth=0, linestyle='-', alpha=0.3,
             dash_capstyle=None,
             hatch=None, # hatchlinewidth=None,
             label='', legend=False,
             zorder=1,
             xrange=None, margins=None, xformat=None, yrange=None, ryrange=None,
             debug=False, **kwarg):

        # Set data.
        # If not specified use self.data.
        if data is None:
            data = self.data

        if indexcol is not None:
            if type(data) == pd.DataFrame:
                # Check if index has already been set for data
                if data.index.name == indexcol:
                    pass

                # Otherwise try to set index to indexcol
                elif indexcol not in data.columns:
                    error = 'indexcol specified as "' + str(indexcol) + '" but not found in data:\n'
                    error += str(data)
                    raise ValueError(error)
                else:
                    data.set_index(indexcol, inplace=True)
                    # If possible, convert to datetime index
                    try:
                        data.index = pd.to_datetime(data.index)
                    except Exception:
                        pass
            else:
                print('WARNING: indexcol specified but data is not DataFrame')

        # Set self.xaxis_type from self.data
        self.xaxis_type = self._set_xaxis_type()
        if debug:
            print('self.xaxis_type = ' + str(self.xaxis_type))
            
        # Set x-axis limits if specified, or use default
        if xrange is not None:
            xrange = self._parse_xrange(xrange, debug=debug)
            self._xrange = xrange
        else:
            xrange = self._xrange
        if debug:
            print('xrange:')
            print(xrange)
            
        # Trim data as needed
        self._trim_data(xrange, debug=debug)
        if debug:
            print(self.data)

        # Check lo and hi are valid columns in self.data
        if lo not in self.data.columns:
            raise RuntimeError('lo specified as "' + str(lo) + '" but not found in data:' + str(data))
        if hi not in self.data.columns:
            raise RuntimeError('hi specified as "' + str(hi) + '" but not found in data:' + str(data))

        if axis == 'left':
            if dash_capstyle:
                entry = self.ax.fill_between(self.data.index, self.data[lo], self.data[hi],
                                             facecolor=color, linewidth=linewidth, edgecolor=linecolor, linestyle=linestyle, alpha=alpha,
                                             hatch=hatch,
                                             zorder=zorder,
                                             dash_capstyle=dash_capstyle)
                # no dash_capstyle specified
            else:
                entry = self.ax.fill_between(self.data.index, self.data[lo], self.data[hi],
                                             facecolor=color, linewidth=linewidth, edgecolor=linecolor, linestyle=linestyle, alpha=alpha,
                                             hatch=hatch,
                                             zorder=zorder)
        else:
            if self.ax_right is None:
                self.ax_right = self.ax.twinx()
                # Set color cycler to be common with the left y-axis.
                # Hack from https://github.com/matplotlib/matplotlib/issues/19479
                self.ax_right._get_lines = self.ax._get_lines

            if dash_capstyle:
                entry = self.ax_right.fill_between(self.data.index, self.data[lo], self.data[hi],
                                                   facecolor=color, linewidth=linewidth, edgecolor=linecolor, linestyle=linestyle, alpha=alpha,
                                                   hatch=hatch,
                                                   zorder=zorder,
                                                   dash_capstyle=dash_capstyle)
                # no dash_capstyle specified
            else:
                entry = self.ax_right.fill_between(self.data.index, self.data[lo], self.data[hi],
                                                   facecolor=color, linewidth=linewidth, edgecolor=linecolor, linestyle=linestyle, alpha=alpha,
                                                   hatch=hatch,
                                                   zorder=zorder)
            
        # If adding to legend
        if legend:
            # Create a Patch that has the same characteristics
            # that can be added to the legend entries.
            entry = Patch(
                facecolor=color,
                edgecolor=linecolor,
                hatch=hatch,
                alpha=alpha,
                label=label)
            self.legend_entries.append(entry)
            self.legend_labels.append(label)

        # Re-create legend
        if self.show_legend:
            self.legend()
            if debug:
                print('called legend()')

        # Set x-axis range if specified
        if xrange is not None or margins is not None:
            if debug:
                print('before calling xrange()')
                print(xrange)
            # If margins is specified, set it for this chart and use it.
            if margins is not None:
                self.margins = margins
            self.xrange(xrange, margins=self.margins, debug=debug)
            if debug:
                print('after calling xrange():')
                print(self._xrange)

        # Set xaxis format.
        # If xformat was specified use it, otherwise use self.xformat
        if xformat is None:
            xformat = self.xformat
        self.xaxis_format(xformat=xformat, debug=debug)
        
        # Set x-axis ticks
        self.ticks(axis='x')

        # Set which x-axis is drawn on top
        self.topaxis(self._topaxis)

        # If yrange is specified use it, otherwise use self.yrange
        if axis == 'left':
            if yrange is None:
                if getattr(self, 'yrange', None) is not None:
                    yrange = self._yrange

            if yrange is not None:
                self._yrange = self._parse_yrange(yrange)
                self.yrange(self._yrange)
        if axis == 'right':
            if ryrange is None:
                if getattr(self, 'ryrange', None) is not None:
                    ryrange = self._ryrange
                    
            if ryrange is not None:
                self._ryrange = self._parse_yrange(ryrange)
                self.ryrange(self._ryrange)
                
    def legend(self, show=True,
               ncol_legend=None, legend_spacing=None,
               legend_left=None, legend_bottom=None, legend_width=None, legend_height=None, legend_mode=None,
               legend_frame=None, legend_color=None, legend_alpha=None, legend_shadow=None, legend_fancybox=None,
               legend_edge=None, legend_linestyle=None, legend_linewidth=None,
               adjust=True,
               legend_header=None, legend_header_color=None, legend_header_fontsize=None):
        '''
        Update legend of Chart.
        Some options can be specified as inputs, if they are provided as None the class  attributes are used.

        If show=False, delete existing legend.
        '''

        if not show:
            # Set internal flag
            self.show_legend = False

            # If a legend already exists, delete
            if self._legend:
                try:
                    self.ax.get_legend().remove()
                    self._legend = None
                except Exception as e:
                    print('WARNING: Could not remove existing legend with exception:')
                    print(e)
                
        else:
            # Set internal flag
            self.show_legend = True

            # Recreate legend if it does not exist.
            # This ensures that any new entries are added in.
            # Use inputs if they are specified, otherwise default to class attributes
            if ncol_legend is None:
                ncol_legend = self.ncol_legend
            if legend_left is None:
                legend_left = self.legend_left
            if legend_bottom is None:
                legend_bottom = self.legend_bottom
            if legend_width is None:
                legend_width = self.legend_width
            if legend_height is None:
                legend_height = self.legend_height
            if legend_mode is None:
                legend_mode = self.legend_mode
            if legend_frame is None:
                legend_frame = self.legend_frame
            if legend_color is None:
                legend_color = self.legend_color
            if legend_alpha is None:
                legend_alpha = self.legend_alpha
            if legend_shadow is None:
                legend_shadow = self.legend_shadow
            if legend_fancybox is None:
                legend_fancybox = self.legend_fancybox
            if legend_edge is None:
                legend_edge = self.legend_edge
            if legend_linestyle is None:
                legend_linestyle = self.legend_linestyle
            if legend_linewidth is None:
                legend_linewidth = self.legend_linewidth
            if legend_spacing is None:
                legend_spacing = self.legend_spacing
            if legend_header is None:
                legend_header = self.legend_header
            if legend_header_color is None:
                legend_header_color = self.legend_header_color
            if legend_header_fontsize is None:
                legend_header_fontsize = self.legend_header_fontsize

            # Save attributes
            self.legend_header = legend_header
    
            # Make sure legend_left + legend_width is less than 1,
            # otherwise we are left with long white margin on right side.
            if adjust:
                if legend_left + legend_width > 1:
                    legend_width = 1 - legend_left
                    self.legend_width = legend_width

            if self.fontname is not None:
                self._legend = self.ax.legend(self.legend_entries, self.legend_labels,
                                              loc='upper left',
                                              labelspacing=legend_spacing,
                                              bbox_transform=self.ax.transAxes,
                                              bbox_to_anchor=(legend_left,legend_bottom,legend_width,legend_height),
                                              mode=legend_mode, borderaxespad=0,
                                              ncol=ncol_legend, fontsize=self.legend_fontsize, frameon=legend_frame,
                                              facecolor=legend_color, framealpha=legend_alpha, shadow=legend_shadow, fancybox=legend_fancybox,
                                              edgecolor=legend_edge,
                                              title=self.legend_header, alignment='left', title_fontsize=self.legend_header_fontsize,
                                              numpoints=1,
                                              prop={"family" : self.fontname}
                                              )
            else:
                self._legend = self.ax.legend(self.legend_entries, self.legend_labels,
                                              loc='upper left',
                                              labelspacing=legend_spacing,
                                              bbox_transform=self.ax.transAxes,
                                              bbox_to_anchor=(legend_left,legend_bottom,legend_width,legend_height),
                                              mode=legend_mode, borderaxespad=0,
                                              ncol=ncol_legend, fontsize=self.legend_fontsize, frameon=legend_frame,
                                              facecolor=legend_color, framealpha=legend_alpha, shadow=legend_shadow, fancybox=legend_fancybox,
                                              edgecolor=legend_edge,
                                              title=self.legend_header, alignment='left', title_fontsize=self.legend_header_fontsize,
                                              numpoints=1
                                              )
                
            # Need to set legend title color
            self._legend.get_title().set_color(self.legend_header_color)
            # Need to set legend edge attributes
            if legend_frame:
                self._legend.get_frame().set_linestyle(legend_linestyle)
                self._legend.get_frame().set_linewidth(legend_linewidth)
        # end of show=True

    def watermark(self, text='Confidential',
                  x=0.5, y=0.5, size=40, font=None, color='gray', alpha=0.3, angle=30, zorder=10):
                  
        '''
        Add watermark.
        Specify position in figure coordinates (0 to 1)
        as well as text to be shown and other attributes such as
        color, size, transparency, angle.
        '''

        # Set up params. If None is given, default to function defaults.
        if x is None:
            x = 0.5
        if y is None:
            y = 0.5
        if size is None:
            size = 40
        if font is None:
            # Use default from rcParams
            matplotlib.rcParams["font.family"]
        if color is None:
            color = 'gray'
        if alpha is None:
            alpha = 0.3
        if angle is None:
            angle = 30
        if zorder is None:
            # This sets the watermark to be in front of everything else.
            # To set behind everything set to 0
            zorder = 10

        self.fig.text(
            x, y,
            text,
            fontsize=size, fontfamily=font, color=color, alpha=alpha,
            ha='center', va='center', rotation=angle, zorder=zorder)
        
    def save(self, filename, dpi=250, bbox_inches='tight', pad_inches=0.02):
        '''
        Save information on chart.
        '''

        # Besides saving chart as file,
        # would be nice to have options to
        # - save data
        # - save style, settings
        # so that chart can be re-generated.

        # Get dirname and create as necessary
        dirname = os.path.dirname(filename)
        if not dirname == '' and not os.path.isdir(dirname):
            try:
                os.makedirs(dirname)
            except Exception as e:
                print('Could not create dir ' + dirname + ' for filename ' + filename)

        try:
            self.fig.savefig(filename, dpi=dpi, bbox_inches=bbox_inches, pad_inches=pad_inches)
        except Exception as e:
            print('WARNING: Could not create file ' + filename + ' with exception:')
            print(e)

    def show(self, debug=False):
        '''
        Show chart.
        '''
        import subprocess

        if debug:
            print('Calling show')

        # The below does not work because due to screen resolution
        # and other factors, the image shown will not match exactly the saved output.
        ### # This is necessary to update charts after they have been shown.
        ### self.fig.canvas.draw()
        ### self.fig.show()

        # To ensure what the user sees is what is saved, save the image as png
        # and show the output.
        tmpfilename = "__preview.png"
        try:
            self.fig.savefig(tmpfilename, dpi=250)
        except Exception as e:
            print('WARNING: Could not save temporary image ' + tmpfilename + ':')
            print(e)
            return

        code = (
            "from PIL import Image; "
            f"Image.open(r'''{tmpfilename}''').show()"
        )
        return subprocess.Popen([sys.executable, "-c", code])        

    def fontlang(self, lang=None):
        '''
        Set up language and corresponding matplotlib.rcParams['font.family']
        '''

        # Set up so that if None is called, this resets
        if lang is None:
            matplotlib.rcParams['font.family'] = 'sans-serif'
        elif str(lang).strip().lower() in ['jp', 'ja']:
            matplotlib.rcParams['font.family'] = 'Yu Gothic'
        elif str(lang).strip().lower() == 'zh':
            matplotlib.rcParams['font.family'] = 'Microsoft YaHei'
        else:
            print('WARNING: lang ' + str(lang) + ' not implemented, defaulting to sans-serif')
            matplotlib.rcParams['font.family'] = 'sans-serif'

    def get_translatable_attrs(self):
        '''
        Create a DataFrame containing all translatable attributes.
        '''
        
        # Get translatable attributes and put into structure
        translatables = OrderedDict()
        translatables['title'] = self.titletext
        translatables['subtitle'] = self.subtitletext
        translatables['xtitle'] = self.xtitletext
        translatables['ytitle'] = self.ytitletext

        # legend
        translatables['legendtitle'] = self.legend_header
        translatables['legend'] = self.legend_labels

        # TODO
        # translatables['caption'] = self.caption
        
        translatables['texts'] = [t.get_text() for t in self.texts]

        # Create a list of attributes
        attrs = []
        for key in translatables:
            if type(translatables[key]) is None:
                attrs.append([key, ''])
            elif type(translatables[key]) == str:
                attrs.append([key, translatables[key]])
            elif type(translatables[key]) == list:
                for iitem, item in enumerate(translatables[key]):
                    attrs.append([key + str(iitem), item])
        # Create DataFrame
        df_attrs = pd.DataFrame(attrs, columns=['key', 'value']).set_index('key')

        return df_attrs

    def exportxl(self, filename, sheet='Sheet 1', lang='EN'):
        '''
        Export the current chart's translatable attributes to an Excel file and sheet.
        If the sheet exists, try to overwrite.
        '''

        # Create DataFrame containing all translatable attributes
        df_attrs = self.get_translatable_attrs()

        write_key_value_excel(
            filename=filename,
            sheetname=sheet,
            df=df_attrs,
            lang=lang,
            alternating_colors=True)

    def importxl(self, filename, sheet='Sheet 1', lang='EN'):
        '''
        Import the current chart's translatable attributes from an Excel file and sheet.
        '''

        from openpyxl import load_workbook

        if not os.path.isfile(filename):
            raise ValueError(f"File {filename} does not exist")

        wb = load_workbook(filename)
        if sheet in wb.sheetnames:
            existing_df = pd.read_excel(filename, sheet_name=sheet, index_col=0)
        else:
            raise ValueError(f"Sheet {sheet} does not exist in {filename}")
        wb.close()

        # language column should always be upper case
        lang = lang.upper()

        # Check if lang exists as a column
        if lang in existing_df.columns:
            df_attrs = existing_df[[lang]]
        else:
            raise ValueError(f"Column {lang} does not exist in file {filename} sheet {sheet}")

        # Set attributes of each element
        for key in df_attrs.index:
            val = df_attrs.loc[key, lang]
            if key == 'title':
                self.title(val)
            elif key == 'subtitle':
                # Remove existing subtitle if it exists
                if self.subtitlehandle:
                    self.subtitlehandle.remove()
                self.subtitle(val)
            elif key == 'xtitle':
                self.xtitle(val)
            elif key == 'ytitle':
                self.ytitle(val)

        # Deal with legend
        if 'legendtitle' in df_attrs.index:
            legendtitle = df_attrs.loc['legendtitle', lang]
            if pd.isnull(legendtitle):
                legendtitle = ''
        # Get all legend entries
        legend_keys = sorted(df_attrs.index[df_attrs.index.str.match(r"^legend\d+$")])
        # Make sure length matches
        if len(self.legend_labels) != len(legend_keys):
            raise ValueError(f"Number of legend entries in chart is {len(self.legend_entries)} while translated input had {len(legend_keys)} entries")
        self.legend_labels = list(df_attrs.loc[legend_keys][lang].values)
        # Recreate legend
        self.legend(legend_header=legendtitle)

        # Deal with texts
        # Get all text entries
        text_keys = sorted(df_attrs.index[df_attrs.index.str.match(r"^texts\d+$")])
        # Make sure length matches
        if len(self.texts) != len(text_keys):
            raise ValueError(f"Number of text entries in chart is {len(self.texts)} while translated input had {len(text_keys)} entries")
        # Modify text
        texts = list(df_attrs.loc[text_keys][lang].values)
        for handle, text in zip(self.texts, texts):
            handle.set_text(text)

def color_bars(barcolors, patches, dataindex, stack=False, debug=False):
    '''
    Color individual bars given a list of dicts of the form
    [{idx1 : color1}, {idx2 : color2}, ...]
    and a list of patches for bars and the data's index.

    Treat patches differently based on whether this is a stacked bar chart or not.
    '''

    if debug:
        print('start of color_bars():')
        print('barcolors:')
        print(barcolors)

    if stack:
        # barcolors should be a list of dicts of form
        # [{'idx1' : 'blue'}, {'idx2' : 'red'}]
        for dict_idx_color in barcolors:
            idx = list(dict_idx_color.keys())[0]
            color = dict_idx_color[idx]
            
            if debug:
                print('idx: ' + str(idx) + ', color = ' + str(color))

            # Find matching index number
            try:
                matching_idx = dataindex.get_loc(idx)
                if debug:
                    print('matching_idx = ' + str(matching_idx))
        
                # Get corresponding patch and change color.
                # If stack=True, there will be 4 patches,
                # positive with, without edge color,
                # negative with, without edge color.
                start = matching_idx.start
                matching_patches = [(ip, p) for ip, p in enumerate(patches) if ip % len(dataindex) == start]
                for ip, p in matching_patches:
                    if debug:
                        hex = mplcolors.to_hex(p.get_facecolor()[:3], keep_alpha=False)
                        print('Filling patch color for ip = ' + str(ip) + ', color = ' + hex)
                        print('x = ' + str(p.get_x()) + ', y = ' + str(p.get_y()) + ', height = ' + str(p.get_height()))
                    p.set_facecolor(color)

                    
            except KeyError:
                print('WARNING: barcolors specified for ' + str(idx) + ' but not found in data')
                continue
        # end of loop over barcolors
    # end of stack

    else:
        # barcolors should be a list of dicts of form
        # [{'idx1' : 'blue'}, {'idx2' : 'red'}]
        for dict_idx_color in barcolors:
            idx = list(dict_idx_color.keys())[0]
            color = dict_idx_color[idx]

            if debug:
                print('idx: ' + str(idx) + ', color = ' + str(color))

            # Find matching index number
            try:
                matching_idx = dataindex.get_loc(idx)
                if debug:
                    print('matching_idx = ' + str(matching_idx))
            
                # Get corresponding patch and change color.
                # If stack=False and no duplicate index values, only one patch should match
                if type(matching_idx) == int:
                    p = patches[matching_idx]
                    if debug:
                        print('matching patches: ' + str(p))
                    p.set_facecolor(color)
                # If there are duplicate index values, matching_idx
                # will be a list of [False, False, True, False, True, ...]
                # where True corresponds to matching indexes.
                elif type(matching_idx) == np.ndarray:
                    for iidx, idx in enumerate(matching_idx):
                        print('iidx = ' + str(iidx) + ' idx = ' + str(idx))
                        if idx:
                            p = patches[iidx]
                            if debug:
                                print('Setting color for patch with iidx = ' + str(iidx))
                                print('x = ' + str(p.get_x()) + ', y = ' + str(p.get_y()) + ', height = ' + str(p.get_height()))
                            p.set_facecolor(color)
                else:
                    raise ValueError('type of matching_idx is ' + str(type(matching_idx)))
            except KeyError:
                print('WARNING: barcolors specified for ' + str(idx) + ' but not found in data')
                continue
        # end of loop over barcolors
    # end of stack=False

def write_key_value_excel(
    filename: str,
    df: pd.DataFrame,
    sheetname: str = 'Sheet1',
    lang: str = "EN",
    alternating_colors: bool = True,
) -> None:
    """
    Utility function used within self.export().
    
    Update or create an Excel sheet containing:
      - a 'key' column
      - one or more language columns, including `lang`

    Behavior
    --------
    - If the Excel file does not exist, create it.
    - If the file exists and the sheet exists:
        * read the sheet into a DataFrame
        * merge/update values for the given `lang` column using `df`
        * preserve any other language columns already present
    - If the file exists but the sheet does not exist, create the sheet.

    Input df requirements
    ---------------------
    - index name must be 'key'
    - must contain exactly one column named 'value'

    Output formatting
    -----------------
    - bold colored header
    - auto-sized columns
    - optional alternating row fills
    - frozen header row
    """
    from pathlib import Path
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    if not isinstance(df, pd.DataFrame):
        raise TypeError("df must be a pandas.DataFrame")

    if df.index.name != "key":
        raise ValueError("df.index.name must be 'key'")

    if list(df.columns) != ["value"]:
        raise ValueError("df must contain exactly one column named 'value'")

    # Make sure lang is upper case
    lang = lang.upper()

    output_path = Path(filename)
    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        raise ValueError('Could not create parent dir for output ' + str(output_path.parent.resolve()))

    # Input data in normalized form: columns ['key', lang]
    new_lang_df = df.reset_index().rename(columns={"value": lang})

    # ------------------------------------------------------------------
    # Read existing sheet if file + sheet already exist
    # ------------------------------------------------------------------
    existing_df = None
    if output_path.exists():
        wb = load_workbook(output_path)
        if sheetname in wb.sheetnames:
            existing_df = pd.read_excel(output_path, sheet_name=sheetname)
        wb.close()

    # ------------------------------------------------------------------
    # Merge/update
    # ------------------------------------------------------------------
    if existing_df is not None:
        if "key" not in existing_df.columns:
            raise ValueError(
                f"Existing sheet '{sheetname}' does not contain a 'key' column."
            )

        merged_df = existing_df.copy()

        # Ensure all keys from new data exist
        merged_df = merged_df.merge(
            new_lang_df[["key"]],
            on="key",
            how="outer",
        )

        # Remove existing lang column if present so we can replace it cleanly
        if lang in merged_df.columns:
            merged_df = merged_df.drop(columns=[lang])

        # Add updated lang column
        merged_df = merged_df.merge(new_lang_df, on="key", how="outer")

        # Put 'key' first, keep all other non-lang columns next, lang last
        other_cols = [c for c in merged_df.columns if c not in ("key", lang)]
        merged_df = merged_df[["key", *other_cols, lang]]
    else:
        merged_df = new_lang_df.copy()

    # ------------------------------------------------------------------
    # Write sheet back into workbook
    # ------------------------------------------------------------------
    mode = "a" if output_path.exists() else "w"

    try:
        with pd.ExcelWriter(
            output_path,
            engine="openpyxl",
            mode=mode,
            if_sheet_exists="replace" if mode == "a" else None,
        ) as writer:
            merged_df.to_excel(writer, sheet_name=sheetname, index=False)
    
            ws = writer.book[sheetname]
    
            # Header styling
            header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
            header_font = Font(bold=True)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
            # Alternating row colors
            if alternating_colors:
                fill_odd = PatternFill(fill_type="solid", fgColor="F7FBFF")
                fill_even = PatternFill(fill_type="solid", fgColor="EAF2F8")
    
                for row_idx in range(2, ws.max_row + 1):
                    fill = fill_odd if (row_idx - 2) % 2 == 0 else fill_even
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = fill
    
            # Alignment
            for row_idx in range(2, ws.max_row + 1):
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                        horizontal="left",
                        vertical="center",
                    )
    
            # Auto-size columns
            for col_idx, column_cells in enumerate(ws.columns, start=1):
                max_len = 0
                for cell in column_cells:
                    value = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, len(value))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)
    
            ws.freeze_panes = "A2"
    except PermissionError:
        raise ValueError('Got PermissionError when trying to write to ' + filename + ', check if Excel file is open') from None
    
def main(argv):
    pass

def read(saved):
    '''
    Read in saved Chart object
    '''

    if not os.path.isfile(saved):
        raise ValueError('File ' + saved + ' does not exist')
        
    # return chart
    pass

if __name__ == '__main__':
    main(sys.argv[1:])
